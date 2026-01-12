"""
排班系统 - 基于 Streamlit 的复杂排班解决方案
支持员工管理、班次管理、排班规则配置和自动排班生成
"""

import streamlit as st
import pandas as pd
import json
import os
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import base64

# 可选导入：用于海报生成和AI功能
try:
    from jinja2 import Template
    HAS_JINJA2 = True
except ImportError:
    HAS_JINJA2 = False

try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False

try:
    from openai import OpenAI
    HAS_OPENAI = True
except ImportError:
    HAS_OPENAI = False

try:
    import difflib
    HAS_DIFFLIB = True
except ImportError:
    HAS_DIFFLIB = False

# 页面配置
st.set_page_config(
    page_title="智能排班系统",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自定义CSS样式 - 全面美化界面
st.markdown("""
<style>
    /* ===== 主题色彩定义 ===== */
    :root {
        --primary-color: #4361ee;
        --primary-light: #7b8cff;
        --primary-dark: #3a0ca3;
        --success-color: #06d6a0;
        --warning-color: #ffd166;
        --danger-color: #ef476f;
        --bg-gradient: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        --sidebar-bg: linear-gradient(180deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    }
    
    /* ===== 隐藏默认元素 ===== */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* ===== 侧边栏深度美化 ===== */
    [data-testid="stSidebar"] {
        background: var(--sidebar-bg);
        box-shadow: 4px 0 15px rgba(0,0,0,0.1);
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3 {
        color: #ffffff !important;
    }
    [data-testid="stSidebar"] .stRadio > label {
        color: rgba(255,255,255,0.9) !important;
        font-size: 13px;
        font-weight: 500;
    }
    [data-testid="stSidebar"] .stRadio [role="radiogroup"] label {
        color: #ffffff !important;
        font-size: 14px;
        padding: 10px 15px;
        border-radius: 10px;
        margin: 3px 0;
        transition: all 0.25s ease;
        border: 1px solid transparent;
    }
    [data-testid="stSidebar"] .stRadio [role="radiogroup"] label:hover {
        background: rgba(255,255,255,0.12);
        border-color: rgba(255,255,255,0.2);
        transform: translateX(3px);
    }
    [data-testid="stSidebar"] .stRadio [role="radiogroup"] label[data-checked="true"] {
        background: linear-gradient(135deg, rgba(102,126,234,0.4) 0%, rgba(118,75,162,0.4) 100%);
        border-color: rgba(255,255,255,0.3);
    }
    [data-testid="stSidebar"] hr {
        border-color: rgba(255,255,255,0.15);
        margin: 15px 0;
    }
    [data-testid="stSidebar"] .stExpander {
        background: rgba(255,255,255,0.05);
        border-radius: 10px;
        border: 1px solid rgba(255,255,255,0.1);
    }
    [data-testid="stSidebar"] .stExpander summary {
        color: #ffffff !important;
    }
    
    /* ===== 主内容区美化 ===== */
    .main .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    
    /* ===== 统计卡片组件 ===== */
    .stat-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 22px 18px;
        border-radius: 16px;
        color: white;
        text-align: center;
        box-shadow: 0 8px 25px rgba(102,126,234,0.3);
        transition: transform 0.3s ease, box-shadow 0.3s ease;
    }
    .stat-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 12px 35px rgba(102,126,234,0.4);
    }
    .stat-card.green {
        background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        box-shadow: 0 8px 25px rgba(17,153,142,0.3);
    }
    .stat-card.green:hover {
        box-shadow: 0 12px 35px rgba(17,153,142,0.4);
    }
    .stat-card.orange {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        box-shadow: 0 8px 25px rgba(245,87,108,0.3);
    }
    .stat-card.orange:hover {
        box-shadow: 0 12px 35px rgba(245,87,108,0.4);
    }
    .stat-card.cyan {
        background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
        box-shadow: 0 8px 25px rgba(79,172,254,0.3);
    }
    .stat-card h2 {
        font-size: 32px;
        margin: 0;
        font-weight: 700;
        color: white !important;
    }
    .stat-card p {
        margin: 8px 0 0 0;
        opacity: 0.92;
        font-size: 14px;
        font-weight: 500;
    }
    
    /* ===== 表格深度美化 ===== */
    .stDataFrame {
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 4px 15px rgba(0,0,0,0.08);
    }
    .stDataFrame [data-testid="stDataFrameResizable"] {
        border-radius: 12px;
    }
    
    /* ===== 按钮组件美化 ===== */
    .stButton > button {
        border-radius: 10px;
        font-weight: 600;
        transition: all 0.25s ease;
        border: none;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(0,0,0,0.15);
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #5a6fd6 0%, #6a4393 100%);
    }
    
    /* ===== 下载按钮美化 ===== */
    .stDownloadButton > button {
        border-radius: 10px;
        font-weight: 500;
        transition: all 0.25s ease;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 18px rgba(0,0,0,0.12);
    }
    
    /* ===== 标题样式 ===== */
    h1 {
        color: #1a1a2e !important;
        font-weight: 800;
        padding-bottom: 12px;
        border-bottom: 4px solid transparent;
        border-image: linear-gradient(90deg, #667eea, #764ba2) 1;
        margin-bottom: 25px;
        font-size: 2rem !important;
    }
    h2 {
        color: #16213e !important;
        font-weight: 700;
        margin-top: 1.5rem;
    }
    h3 {
        color: #2d3748 !important;
        font-weight: 600;
    }
    
    /* ===== 信息提示框美化 ===== */
    .stAlert {
        border-radius: 12px;
        border: none;
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
    }
    
    /* ===== 输入框美化 ===== */
    .stTextInput > div > div > input,
    .stSelectbox > div > div,
    .stMultiSelect > div > div {
        border-radius: 10px;
        border: 2px solid #e2e8f0;
        transition: border-color 0.2s ease, box-shadow 0.2s ease;
    }
    .stTextInput > div > div > input:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 3px rgba(102,126,234,0.15);
    }
    
    /* ===== 展开面板美化 ===== */
    .streamlit-expanderHeader {
        border-radius: 10px;
        background: #f8fafc;
        font-weight: 600;
    }
    .streamlit-expanderContent {
        border-radius: 0 0 10px 10px;
    }
    
    /* ===== 班次标签样式 ===== */
    .shift-tag {
        display: inline-block;
        padding: 6px 14px;
        border-radius: 25px;
        font-size: 13px;
        font-weight: 600;
        margin: 3px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    .shift-early {
        background: linear-gradient(135deg, #ffecd2 0%, #fcb69f 100%);
        color: #c44536;
    }
    .shift-early-early {
        background: linear-gradient(135deg, #fff1eb 0%, #ace0f9 100%);
        color: #1a535c;
    }
    .shift-late {
        background: linear-gradient(135deg, #e0c3fc 0%, #8ec5fc 100%);
        color: #3d348b;
    }
    .shift-rest {
        background: linear-gradient(135deg, #d4fc79 0%, #96e6a1 100%);
        color: #2d6a4f;
    }
    .shift-standby {
        background: linear-gradient(135deg, #ffeaa7 0%, #dfe6e9 100%);
        color: #6c5ce7;
    }
    
    /* ===== 流程步骤指示器 ===== */
    .workflow-step {
        display: inline-flex;
        align-items: center;
        padding: 8px 16px;
        margin: 5px;
        border-radius: 25px;
        font-size: 13px;
        font-weight: 500;
        background: #f1f5f9;
        color: #64748b;
        transition: all 0.2s ease;
    }
    .workflow-step.active {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        box-shadow: 0 4px 15px rgba(102,126,234,0.4);
    }
    .workflow-step.completed {
        background: linear-gradient(135deg, #06d6a0 0%, #1b9aaa 100%);
        color: white;
    }
    
    /* ===== 图例样式 ===== */
    .legend-item {
        display: inline-flex;
        align-items: center;
        margin: 5px 12px 5px 0;
        font-size: 13px;
    }
    .legend-color {
        width: 16px;
        height: 16px;
        border-radius: 4px;
        margin-right: 6px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    
    /* ===== 分割线美化 ===== */
    hr {
        border: none;
        height: 2px;
        background: linear-gradient(90deg, transparent, #e2e8f0, transparent);
        margin: 25px 0;
    }
    
    /* ===== 标签页美化 ===== */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 10px 10px 0 0;
        padding: 12px 24px;
        font-weight: 600;
    }
    
    /* ===== 进度条美化 ===== */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #667eea, #764ba2);
        border-radius: 10px;
    }
    
    /* ========================================
       ===== 移动端响应式设计 (Mobile) =====
       ======================================== */
    
    /* ----- 平板设备 (768px - 1024px) ----- */
    @media screen and (max-width: 1024px) {
        .main .block-container {
            padding: 1rem 1.5rem;
        }
        
        .stat-card h2 {
            font-size: 26px;
        }
        
        .stat-card p {
            font-size: 12px;
        }
        
        h1 {
            font-size: 1.6rem !important;
        }
        
        h2 {
            font-size: 1.3rem !important;
        }
    }
    
    /* ----- 手机设备 (小于 768px) ----- */
    @media screen and (max-width: 768px) {
        /* 主内容区适配 */
        .main .block-container {
            padding: 0.8rem 1rem !important;
            max-width: 100% !important;
        }
        
        /* 标题缩小 */
        h1 {
            font-size: 1.4rem !important;
            padding-bottom: 8px;
            margin-bottom: 15px;
        }
        
        h2 {
            font-size: 1.15rem !important;
            margin-top: 1rem;
        }
        
        h3 {
            font-size: 1rem !important;
        }
        
        /* 侧边栏优化 */
        [data-testid="stSidebar"] {
            min-width: 240px !important;
            width: 240px !important;
        }
        
        [data-testid="stSidebar"] .stRadio [role="radiogroup"] label {
            font-size: 13px;
            padding: 12px 10px;
        }
        
        /* 统计卡片：手机上竖向排列 */
        .stat-card {
            padding: 16px 12px;
            border-radius: 12px;
            margin-bottom: 10px;
        }
        
        .stat-card h2 {
            font-size: 24px;
        }
        
        .stat-card p {
            font-size: 11px;
            margin-top: 4px;
        }
        
        /* 按钮适配：增大触摸区域 */
        .stButton > button {
            min-height: 48px !important;
            font-size: 14px !important;
            padding: 12px 16px !important;
            width: 100% !important;
        }
        
        .stDownloadButton > button {
            min-height: 48px !important;
            font-size: 14px !important;
            padding: 12px 16px !important;
            width: 100% !important;
        }
        
        /* 输入框适配 */
        .stTextInput > div > div > input {
            min-height: 48px !important;
            font-size: 16px !important;
            padding: 12px !important;
        }
        
        .stSelectbox > div > div {
            min-height: 48px !important;
        }
        
        .stSelectbox > div > div > div {
            font-size: 14px !important;
        }
        
        .stNumberInput > div > div > input {
            min-height: 48px !important;
            font-size: 16px !important;
        }
        
        /* 表格容器：横向滚动 */
        .stDataFrame {
            overflow-x: auto !important;
            -webkit-overflow-scrolling: touch;
        }
        
        .stDataFrame > div {
            overflow-x: auto !important;
        }
        
        /* 展开面板适配 */
        .streamlit-expanderHeader {
            font-size: 14px;
            padding: 12px;
        }
        
        /* 图例适配 */
        .legend-item {
            font-size: 11px;
            margin: 4px 8px 4px 0;
        }
        
        .legend-color {
            width: 14px;
            height: 14px;
        }
        
        /* 班次标签适配 */
        .shift-tag {
            padding: 5px 10px;
            font-size: 11px;
            margin: 2px;
        }
        
        /* 信息提示框 */
        .stAlert {
            padding: 12px;
            font-size: 13px;
        }
        
        /* 分割线 */
        hr {
            margin: 15px 0;
        }
        
        /* 标签页适配 */
        .stTabs [data-baseweb="tab"] {
            padding: 10px 12px;
            font-size: 13px;
        }
    }
    
    /* ----- 超小屏幕设备 (小于 480px) ----- */
    @media screen and (max-width: 480px) {
        .main .block-container {
            padding: 0.5rem 0.8rem !important;
        }
        
        h1 {
            font-size: 1.2rem !important;
        }
        
        h2 {
            font-size: 1.05rem !important;
        }
        
        .stat-card {
            padding: 14px 10px;
        }
        
        .stat-card h2 {
            font-size: 20px;
        }
        
        .stat-card p {
            font-size: 10px;
        }
        
        /* 侧边栏更窄 */
        [data-testid="stSidebar"] {
            min-width: 220px !important;
            width: 220px !important;
        }
        
        [data-testid="stSidebar"] .stRadio [role="radiogroup"] label {
            font-size: 12px;
            padding: 10px 8px;
        }
        
        /* 表格字体缩小 */
        .stDataFrame {
            font-size: 11px !important;
        }
    }
    
    /* ----- 横屏模式优化 ----- */
    @media screen and (max-width: 768px) and (orientation: landscape) {
        .main .block-container {
            padding: 0.5rem 1rem !important;
        }
        
        h1 {
            font-size: 1.3rem !important;
            margin-bottom: 10px;
        }
        
        .stat-card {
            padding: 12px 10px;
        }
        
        .stat-card h2 {
            font-size: 22px;
        }
    }
    
    /* ----- 触摸设备优化 ----- */
    @media (hover: none) and (pointer: coarse) {
        /* 增大所有可点击元素的触摸区域 */
        .stButton > button,
        .stDownloadButton > button {
            min-height: 48px !important;
        }
        
        /* 禁用 hover 效果（触摸设备上不需要） */
        .stButton > button:hover {
            transform: none;
        }
        
        .stat-card:hover {
            transform: none;
        }
        
        /* 增大复选框和单选框的触摸区域 */
        .stCheckbox > label,
        .stRadio > label {
            padding: 12px 8px !important;
        }
    }
</style>
""", unsafe_allow_html=True)

# 数据存储目录
DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)

# 数据文件路径
EMPLOYEES_FILE = os.path.join(DATA_DIR, "employees.json")
SHIFTS_FILE = os.path.join(DATA_DIR, "shifts.json")
RULES_FILE = os.path.join(DATA_DIR, "rules.json")
SCHEDULE_FILE = os.path.join(DATA_DIR, "schedule.json")


def load_json(filepath: str, default: dict = None) -> dict:
    """加载 JSON 文件"""
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    return default if default is not None else {}


def save_json(filepath: str, data: dict):
    """保存 JSON 文件"""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def init_session_state():
    """初始化会话状态"""
    if 'employees' not in st.session_state:
        st.session_state.employees = load_json(EMPLOYEES_FILE, default={})
    if 'shifts' not in st.session_state:
        st.session_state.shifts = load_json(SHIFTS_FILE, default={})
    if 'rules' not in st.session_state:
        st.session_state.rules = load_json(RULES_FILE, default={
            "min_rest_hours": 8,
            "max_consecutive_days": 5,
            "min_weekly_hours": 30,
            "max_weekly_hours": 50,
            "preferred_shifts": {},
            "unavailable_days": {}
        })
    if 'schedule' not in st.session_state:
        st.session_state.schedule = load_json(SCHEDULE_FILE, default={})
    if 'schedule_period' not in st.session_state:
        st.session_state.schedule_period = {
            "start_date": datetime.now().strftime("%Y-%m-%d"),
            "end_date": (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
        }


def employee_management():
    """员工管理页面"""
    st.header("👥 员工管理")
    
    employees = st.session_state.employees
    
    # 顶部统计卡片
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("👥 员工总数", len(employees), delta=None)
    with col2:
        skills_count = sum(len(emp.get("skills", [])) for emp in employees.values())
        st.metric("🎯 技能总数", skills_count)
    with col3:
        rest_days = len([e for e in employees.values() if e.get("rest_day")])
        st.metric("📅 已设休息日", f"{rest_days}/{len(employees)}")
    
    st.markdown("---")
    
    if not employees:
        st.info("💡 暂无员工，请在下方表格中添加员工信息")
        # 创建空表格供编辑
        empty_df = pd.DataFrame({
            "ID": [""],
            "姓名": [""],
            "职位": [""],
            "技能": [""],
            "每周工作小时": [40],
            "休息日": [""],
            "偏好班次": [""],
            "不可用日期": [""]
        })
        edited_df = st.data_editor(
            empty_df,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "ID": st.column_config.TextColumn("ID *", required=True, help="员工唯一标识"),
                "姓名": st.column_config.TextColumn("姓名 *", required=True),
                "职位": st.column_config.TextColumn("职位"),
                "技能": st.column_config.TextColumn("技能", help="多个技能用逗号分隔，如：收银, 备餐"),
                "每周工作小时": st.column_config.NumberColumn("每周工作小时", min_value=0, max_value=80, default=40),
                "休息日": st.column_config.TextColumn("休息日", help="如：周一、周二等"),
                "偏好班次": st.column_config.TextColumn("偏好班次", help="多个班次用逗号分隔，如：早班, 晚班"),
                "不可用日期": st.column_config.TextColumn("不可用日期", help="格式：YYYY-MM-DD，多个日期用逗号分隔")
            }
        )
        
        if st.button("💾 保存修改", type="primary", use_container_width=True):
            new_employees = {}
            for _, row in edited_df.iterrows():
                emp_id = str(row["ID"]).strip()
                name = str(row["姓名"]).strip()
                if emp_id and name:
                    skills = [s.strip() for s in str(row["技能"]).split(",") if s.strip()] if pd.notna(row["技能"]) else []
                    preferred_shifts = [s.strip() for s in str(row["偏好班次"]).split(",") if s.strip()] if pd.notna(row["偏好班次"]) else []
                    
                    new_employees[emp_id] = {
                        "name": name,
                        "position": str(row["职位"]).strip() if pd.notna(row["职位"]) else "",
                        "skills": skills,
                        "weekly_hours": int(row["每周工作小时"]) if pd.notna(row["每周工作小时"]) else 40,
                        "rest_day": str(row["休息日"]).strip() if pd.notna(row["休息日"]) else "",
                        "preferred_shifts": preferred_shifts,
                        "unavailable_days": str(row["不可用日期"]).strip() if pd.notna(row["不可用日期"]) else ""
                    }
            
            if new_employees:
                save_json(EMPLOYEES_FILE, new_employees)
                st.session_state.employees = new_employees
                st.success("✅ 员工信息已保存")
                st.rerun()
            else:
                st.warning("⚠️ 请至少添加一个有效的员工（需要ID和姓名）")
        return
    
    # 构建可编辑的数据框
    df_data = []
    for emp_id, emp in employees.items():
        df_data.append({
            "ID": emp_id,
            "姓名": emp.get("name", ""),
            "职位": emp.get("position", ""),
            "技能": ", ".join(emp.get("skills", [])),
            "每周工作小时": emp.get("weekly_hours", 40),
            "休息日": emp.get("rest_day", ""),
            "偏好班次": ", ".join(emp.get("preferred_shifts", [])),
            "不可用日期": emp.get("unavailable_days", "")
        })
    
    df = pd.DataFrame(df_data)
    
    # 使用可编辑表格
    st.markdown("**可直接在表格中编辑，修改完成后点击下方“保存修改”按钮**")
    edited_df = st.data_editor(
        df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "ID": st.column_config.TextColumn("ID *", required=True, help="员工唯一标识"),
            "姓名": st.column_config.TextColumn("姓名 *", required=True),
            "职位": st.column_config.TextColumn("职位"),
            "技能": st.column_config.TextColumn("技能", help="多个技能用逗号分隔"),
            "每周工作小时": st.column_config.NumberColumn("每周工作小时", min_value=0, max_value=80),
            "休息日": st.column_config.TextColumn("休息日", help="如：周一、周二等"),
            "偏好班次": st.column_config.TextColumn("偏好班次", help="多个班次用逗号分隔"),
            "不可用日期": st.column_config.TextColumn("不可用日期", help="格式：YYYY-MM-DD，多个用逗号分隔")
        }
    )
    
    col_save, col_delete, col_export1, col_export2 = st.columns([2, 2, 2, 2])
    
    with col_save:
        if st.button("💾 保存修改", type="primary", use_container_width=True):
            new_employees = {}
            for _, row in edited_df.iterrows():
                emp_id = str(row["ID"]).strip()
                name = str(row["姓名"]).strip()
                if emp_id and name:
                    skills = [s.strip() for s in str(row["技能"]).split(",") if s.strip()] if pd.notna(row["技能"]) and str(row["技能"]).strip() else []
                    preferred_shifts = [s.strip() for s in str(row["偏好班次"]).split(",") if s.strip()] if pd.notna(row["偏好班次"]) and str(row["偏好班次"]).strip() else []
                    
                    new_employees[emp_id] = {
                        "name": name,
                        "position": str(row["职位"]).strip() if pd.notna(row["职位"]) else "",
                        "skills": skills,
                        "weekly_hours": int(row["每周工作小时"]) if pd.notna(row["每周工作小时"]) else 40,
                        "rest_day": str(row["休息日"]).strip() if pd.notna(row["休息日"]) else "",
                        "preferred_shifts": preferred_shifts,
                        "unavailable_days": str(row["不可用日期"]).strip() if pd.notna(row["不可用日期"]) else ""
                    }
            
            if new_employees:
                save_json(EMPLOYEES_FILE, new_employees)
                st.session_state.employees = new_employees
                st.success("✅ 员工信息已保存")
                st.rerun()
            else:
                st.warning("⚠️ 至少需要一个有效的员工（需要ID和姓名）")
    
    with col_delete:
        if st.button("🗑️ 删除选中行", use_container_width=True):
            st.info("💡 要删除员工，请先清空该行的ID或姓名，然后点击“保存修改”")
    
    with col_export1:
        csv = df.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            label="📥 导出 CSV",
            data=csv,
            file_name=f"员工列表_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True
        )
    
    with col_export2:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='员工列表')
        excel_data = output.getvalue()
        st.download_button(
            label="📥 导出 Excel",
            data=excel_data,
            file_name=f"员工列表_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )


def shift_management():
    """班次管理页面"""
    st.header("⏰ 班次管理")
    
    shifts = st.session_state.shifts
    
    # 顶部统计
    col_stat1, col_stat2, col_stat3 = st.columns(3)
    with col_stat1:
        st.metric("📋 班次数量", len(shifts))
    with col_stat2:
        total_staff = sum(s.get("required_staff", 1) for s in shifts.values())
        st.metric("👥 每日需人", total_staff)
    with col_stat3:
        total_hours = sum(s.get("duration_hours", 8) for s in shifts.values())
        st.metric("⏱️ 总工时", f"{total_hours}h")
    
    st.markdown("---")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("班次列表")
        shifts = st.session_state.shifts
        
        if shifts:
            df = pd.DataFrame([
                {
                    "班次名称": shift_id,
                    "开始时间": shift.get("start_time", ""),
                    "结束时间": shift.get("end_time", ""),
                    "所需人数": shift.get("required_staff", 1),
                    "所需技能": ", ".join(shift.get("required_skills", [])),
                    "持续时间（小时）": shift.get("duration_hours", 8)
                }
                for shift_id, shift in shifts.items()
            ])
            st.dataframe(df, use_container_width=True, hide_index=True)
            
            # 导出班次列表
            st.markdown("### 导出班次列表")
            col_csv, col_excel = st.columns(2)
            with col_csv:
                csv = df.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    label="📥 导出为 CSV",
                    data=csv,
                    file_name=f"班次列表_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            with col_excel:
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='班次列表')
                excel_data = output.getvalue()
                st.download_button(
                    label="📥 导出为 Excel",
                    data=excel_data,
                    file_name=f"班次列表_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.info("暂无班次，请添加班次信息")
    
    with col2:
        st.subheader("添加/编辑班次")
        with st.form("shift_form"):
            shift_id = st.text_input("班次名称 *", placeholder="例如: 早班")
            start_time = st.time_input("开始时间 *", value=datetime.strptime("09:00", "%H:%M").time())
            end_time = st.time_input("结束时间 *", value=datetime.strptime("17:00", "%H:%M").time())
            required_staff = st.number_input("所需人数 *", min_value=1, value=2, step=1)
            required_skills_input = st.text_input("所需技能（逗号分隔）", placeholder="例如: 收银, 备餐")
            
            submitted = st.form_submit_button("保存", use_container_width=True)
            
            if submitted:
                if not shift_id:
                    st.error("班次名称为必填项")
                else:
                    # 计算持续时间
                    start_dt = datetime.combine(datetime.today(), start_time)
                    end_dt = datetime.combine(datetime.today(), end_time)
                    if end_dt <= start_dt:
                        end_dt += timedelta(days=1)  # 跨天情况
                    duration = (end_dt - start_dt).total_seconds() / 3600
                    
                    required_skills = [s.strip() for s in required_skills_input.split(",") if s.strip()] if required_skills_input else []
                    
                    shifts[shift_id] = {
                        "start_time": start_time.strftime("%H:%M"),
                        "end_time": end_time.strftime("%H:%M"),
                        "required_staff": int(required_staff),
                        "required_skills": required_skills,
                        "duration_hours": round(duration, 2)
                    }
                    save_json(SHIFTS_FILE, shifts)
                    st.session_state.shifts = shifts
                    st.success(f"班次 {shift_id} 已保存")
                    st.rerun()
    
    # 删除班次
    if shifts:
        st.subheader("删除班次")
        shift_ids = list(shifts.keys())
        selected_shift_id = st.selectbox("选择要删除的班次", shift_ids, key="delete_shift")
        if st.button("删除班次", type="primary"):
            del shifts[selected_shift_id]
            save_json(SHIFTS_FILE, shifts)
            st.session_state.shifts = shifts
            st.success(f"班次 {selected_shift_id} 已删除")
            st.rerun()


def rules_management():
    """排班规则管理页面"""
    st.header("⚙️ 排班规则配置")
    
    rules = st.session_state.rules
    employees = st.session_state.employees
    schedule = st.session_state.schedule
    
    # 显示当前规则统计
    st.subheader("📊 当前规则统计")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("最小休息时间", f"{rules.get('min_rest_hours', 8)} 小时")
    with col2:
        st.metric("最大连续工作天数", f"{rules.get('max_consecutive_days', 5)} 天")
    with col3:
        st.metric("每周最少工作小时", f"{rules.get('min_weekly_hours', 30)} 小时")
    with col4:
        st.metric("每周最多工作小时", f"{rules.get('max_weekly_hours', 50)} 小时")
    
    # 规则效果预览（如果有排班数据）
    if schedule and employees:
        st.subheader("📈 规则效果预览")
        
        # 检查连续工作天数
        consecutive_violations = []
        for emp_id, emp in employees.items():
            emp_name = emp.get("name", emp_id)
            dates = sorted(schedule.keys())
            consecutive_days = 0
            max_consecutive = 0
            max_rest_hours = rules.get("max_consecutive_days", 5)
            
            for i, date_str in enumerate(dates):
                date_schedule = schedule[date_str]
                assignments = date_schedule.get("assignments", {})
                if emp_id in assignments:
                    consecutive_days += 1
                    max_consecutive = max(max_consecutive, consecutive_days)
                else:
                    consecutive_days = 0
            
            if max_consecutive > max_rest_hours:
                consecutive_violations.append({
                    "员工": emp_name,
                    "最大连续工作天数": max_consecutive,
                    "规则限制": max_rest_hours,
                    "状态": "⚠️ 超限"
                })
        
        # 检查每周工作小时
        weekly_hours_violations = []
        if len(schedule) >= 7:
            dates = sorted(schedule.keys())
            weeks = len(dates) // 7
            
            for emp_id, emp in employees.items():
                emp_name = emp.get("name", emp_id)
                weekly_hours_target = emp.get("weekly_hours", 40)
                min_weekly_hours = rules.get("min_weekly_hours", 30)
                max_weekly_hours = rules.get("max_weekly_hours", 50)
                
                # 计算每周工作小时
                for week_idx in range(weeks):
                    week_dates = dates[week_idx * 7:(week_idx + 1) * 7]
                    week_hours = 0
                    
                    for date_str in week_dates:
                        date_schedule = schedule[date_str]
                        assignments = date_schedule.get("assignments", {})
                        if emp_id in assignments:
                            shift_id = assignments[emp_id]
                            shift = st.session_state.shifts.get(shift_id, {})
                            week_hours += shift.get("duration_hours", 8)
                    
                    if week_hours < min_weekly_hours:
                        weekly_hours_violations.append({
                            "员工": emp_name,
                            "周次": week_idx + 1,
                            "实际工作小时": round(week_hours, 1),
                            "最低要求": min_weekly_hours,
                            "状态": "⚠️ 不足"
                        })
                    elif week_hours > max_weekly_hours:
                        weekly_hours_violations.append({
                            "员工": emp_name,
                            "周次": week_idx + 1,
                            "实际工作小时": round(week_hours, 1),
                            "最高限制": max_weekly_hours,
                            "状态": "⚠️ 超限"
                        })
        
        # 显示违规情况
        if consecutive_violations:
            st.warning(f"⚠️ 发现 {len(consecutive_violations)} 个连续工作天数违规")
            violations_df = pd.DataFrame(consecutive_violations)
            st.dataframe(violations_df, use_container_width=True, hide_index=True)
        else:
            st.success("✅ 连续工作天数符合规则")
        
        if weekly_hours_violations:
            st.warning(f"⚠️ 发现 {len(weekly_hours_violations)} 个每周工作小时违规")
            violations_df = pd.DataFrame(weekly_hours_violations)
            st.dataframe(violations_df, use_container_width=True, hide_index=True)
        else:
            st.success("✅ 每周工作小时符合规则")
    
    # 规则配置
    st.subheader("⚙️ 规则配置")
    
    with st.form("rules_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### 基本规则")
            min_rest_hours = st.number_input(
                "最小休息时间（小时）",
                min_value=0,
                max_value=24,
                value=rules.get("min_rest_hours", 8),
                step=1,
                help="两次班次之间的最小休息时间"
            )
            max_consecutive_days = st.number_input(
                "最大连续工作天数",
                min_value=1,
                max_value=10,
                value=rules.get("max_consecutive_days", 5),
                step=1,
                help="员工可以连续工作的最大天数"
            )
        
        with col2:
            st.markdown("#### 工作时间规则")
            min_weekly_hours = st.number_input(
                "每周最少工作小时",
                min_value=0,
                max_value=80,
                value=rules.get("min_weekly_hours", 30),
                step=1,
                help="员工每周最少需要工作的小时数"
            )
            max_weekly_hours = st.number_input(
                "每周最多工作小时",
                min_value=0,
                max_value=80,
                value=rules.get("max_weekly_hours", 50),
                step=1,
                help="员工每周最多可以工作的小时数"
            )
        
        # 规则验证
        if min_weekly_hours > max_weekly_hours:
            st.error("❌ 每周最少工作小时不能大于每周最多工作小时")
            form_valid = False
        else:
            form_valid = True
        
        submitted = st.form_submit_button("💾 保存规则", use_container_width=True, type="primary")
        
        if submitted and form_valid:
            rules.update({
                "min_rest_hours": int(min_rest_hours),
                "max_consecutive_days": int(max_consecutive_days),
                "min_weekly_hours": int(min_weekly_hours),
                "max_weekly_hours": int(max_weekly_hours)
            })
            save_json(RULES_FILE, rules)
            st.session_state.rules = rules
            st.success("✅ 规则已保存")
            st.rerun()
    
    # 规则说明
    with st.expander("📖 规则说明"):
        st.markdown("""
        **基本规则**：
        - **最小休息时间**：员工两次班次之间必须休息的最少小时数，确保员工有足够的休息时间
        - **最大连续工作天数**：员工可以连续工作的最大天数，超过此天数必须安排休息
        
        **工作时间规则**：
        - **每周最少工作小时**：员工每周最少需要工作的小时数，确保员工有足够的工作量
        - **每周最多工作小时**：员工每周最多可以工作的小时数，防止过度工作
        
        **注意事项**：
        - 修改规则后，建议重新生成排班表以确保符合新规则
        - 如果现有排班表违反规则，会在"规则效果预览"中显示
        """)


def get_weekday_chinese(date_str: str) -> str:
    """获取日期的中文星期几"""
    weekday_map = {
        0: "周一",
        1: "周二",
        2: "周三",
        3: "周四",
        4: "周五",
        5: "周六",
        6: "周日"
    }
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    return weekday_map[date_obj.weekday()]


def check_conflicts(emp_id: str, shift_id: str, date: str, schedule: dict) -> List[str]:
    """检查排班冲突（只检查日期，不考虑工作时间）"""
    conflicts = []
    employees = st.session_state.employees
    shifts = st.session_state.shifts
    
    if emp_id not in employees:
        conflicts.append(f"员工 {emp_id} 不存在")
        return conflicts
    
    if shift_id not in shifts:
        conflicts.append(f"班次 {shift_id} 不存在")
        return conflicts
    
    emp = employees[emp_id]
    shift = shifts[shift_id]
    
    # 检查中文星期几的休息日
    weekday_chinese = get_weekday_chinese(date)
    rest_day = emp.get("rest_day", "")
    if rest_day and weekday_chinese == rest_day:
        conflicts.append(f"员工 {emp.get('name', emp_id)} 在 {weekday_chinese} 休息")
    
    # 检查不可用日期（日期字符串格式）
    unavailable_days = emp.get("unavailable_days", "")
    if unavailable_days and date in unavailable_days:
        conflicts.append(f"员工 {emp.get('name', emp_id)} 在 {date} 不可用")
    
    # 检查技能匹配
    required_skills = shift.get("required_skills", [])
    emp_skills = emp.get("skills", [])
    if required_skills and not any(skill in emp_skills for skill in required_skills):
        conflicts.append(f"员工 {emp.get('name', emp_id)} 缺乏所需技能: {', '.join(required_skills)}")
    
    # 检查同一天是否已有班次
    if date in schedule:
        if emp_id in schedule[date].get("assignments", {}):
            conflicts.append(f"员工 {emp.get('name', emp_id)} 在 {date} 已有其他班次")
    
    return conflicts


def generate_schedule():
    """自动生成排班表"""
    st.header("🤖 生成排班表")
    
    employees = st.session_state.employees
    shifts = st.session_state.shifts
    
    # 检查前置条件
    if not employees or not shifts:
        st.warning("⚠️ 请先完成以下步骤：")
        col1, col2 = st.columns(2)
        with col1:
            if not employees:
                st.error("❌ 员工管理：尚未添加员工")
            else:
                st.success(f"✅ 员工管理：已有 {len(employees)} 名员工")
        with col2:
            if not shifts:
                st.error("❌ 班次管理：尚未添加班次")
            else:
                st.success(f"✅ 班次管理：已有 {len(shifts)} 个班次")
        return
    
    # 顶部统计
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                padding: 20px; border-radius: 12px; color: white; margin-bottom: 20px;">
        <div style="display: flex; justify-content: space-around; text-align: center;">
            <div><h2 style="margin:0; color:white;">{len(employees)}</h2><p style="margin:0; opacity:0.9;">员工</p></div>
            <div><h2 style="margin:0; color:white;">{len(shifts)}</h2><p style="margin:0; opacity:0.9;">班次</p></div>
            <div><h2 style="margin:0; color:white;">{sum(s.get('required_staff',1) for s in shifts.values())}</h2><p style="margin:0; opacity:0.9;">每日需人</p></div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 日期选择
    col1, col2, col3 = st.columns([2, 2, 1])
    
    with col1:
        start_date = st.date_input(
            "📅 开始日期",
            value=datetime.strptime(st.session_state.schedule_period["start_date"], "%Y-%m-%d").date(),
            help="排班表开始的日期"
        )
    
    with col2:
        end_date = st.date_input(
            "📅 结束日期",
            value=datetime.strptime(st.session_state.schedule_period["end_date"], "%Y-%m-%d").date(),
            help="排班表结束的日期"
        )
    
    with col3:
        st.write("")  # 占位
        st.write("")  # 占位
    
    if start_date >= end_date:
        st.error("❌ 结束日期必须晚于开始日期")
        return
    
    rules = st.session_state.rules
    days_count = (end_date - start_date).days + 1
    
    st.info(f"📊 将生成 **{days_count} 天** 的排班表")
    
    if st.button("🚀 生成排班表", type="primary", use_container_width=True):
        schedule = {}
        date_range = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]
        
        # 初始化排班表
        for date in date_range:
            date_str = date.strftime("%Y-%m-%d")
            schedule[date_str] = {
                "assignments": {},
                "shift_counts": {shift_id: 0 for shift_id in shifts.keys()}
            }
        
        # 获取特殊规则
        special_rules = rules.get("special_rules", {})
        fixed_early_early_employees = special_rules.get("fixed_early_early_shift_employees", [])
        monday_no_early_early = special_rules.get("monday_no_early_early_shift", False)
        
        employee_list = list(employees.keys())
        shift_list = list(shifts.keys())
        
        # 处理固定早早班人员
        fixed_early_early_shift = "二期水吧-早早班"
        
        # 特殊班次配置：允许为空的班次
        # 注意：二期水吧-早班的所需人数由动态规则决定（根据早早班的排班情况）
        flexible_shifts = {
            "二期水吧-早班": {
                "allow_empty": True   # 允许为空（找不到人可以跳过）
                # required_staff 由动态规则决定：如果早早班没人则需要2人，有人则只需1人
            }
        }
        
        # 初始化员工工作量跟踪
        employee_workload = {emp_id: {"days": 0, "hours": 0} for emp_id in employee_list}
        
        # 初始化员工休息周期班次跟踪
        # 结构: {emp_id: {"current_type": "早班", "last_work_date": "2024-01-01", "rested": False}}
        # current_type: 当前休息周期内的班次类型
        # last_work_date: 上次工作日期
        # rested: 是否刚休息过（用于判断是否需要轮换）
        employee_shift_cycle = {emp_id: {"current_type": None, "last_work_date": None, "rested": False} for emp_id in employee_list}
        
        def get_shift_type(shift_id: str) -> str:
            """获取班次类型（早早班/早班/晚班/其他）"""
            if "早早" in shift_id:
                return "早早班"
            elif "早" in shift_id:
                return "早班"
            elif "晚" in shift_id:
                return "晚班"
            elif "中" in shift_id:
                return "中班"
            else:
                return "其他"
        
        def get_opposite_shift_type(shift_type: str) -> str:
            """获取轮换后的班次类型（早班<->晚班）"""
            if shift_type == "早班":
                return "晚班"
            elif shift_type == "晚班":
                return "早班"
            else:
                return shift_type  # 其他类型不轮换
        
        def check_if_rested(emp_id: str, current_date_str: str) -> bool:
            """检查员工是否刚经过固定休息日（只有固定休息日才算真正休息，无岗是待岗不算休息）"""
            cycle_info = employee_shift_cycle.get(emp_id, {})
            last_work = cycle_info.get("last_work_date")
            
            if not last_work:
                return False  # 第一次排班，不算休息过
            
            emp = employees.get(emp_id, {})
            rest_day = emp.get("rest_day", "")  # 员工的固定休息日，如"周一"
            
            if not rest_day:
                return False  # 没有设置固定休息日，不会轮换
            
            # 检查从上次工作日到今天之间是否经过了固定休息日
            last_date = datetime.strptime(last_work, "%Y-%m-%d")
            current_date = datetime.strptime(current_date_str, "%Y-%m-%d")
            
            weekday_map = {0: "周一", 1: "周二", 2: "周三", 3: "周四", 4: "周五", 5: "周六", 6: "周日"}
            
            # 遍历从上次工作日+1到今天的所有日期，看是否包含固定休息日
            check_date = last_date + timedelta(days=1)
            while check_date <= current_date:
                if weekday_map[check_date.weekday()] == rest_day:
                    return True  # 经过了固定休息日
                check_date += timedelta(days=1)
            
            return False  # 没有经过固定休息日
        
        def get_employee_score(emp_id: str, shift_id: str, date_str: str, current_workload: dict) -> float:
            """计算员工分配优先级分数（分数越低优先级越高）"""
            emp = employees[emp_id]
            shift = shifts[shift_id]
            
            score = 0.0
            
            # 工作量平衡：工作天数少的优先
            score += current_workload[emp_id]["days"] * 10
            
            # 技能匹配：有技能的优先（减少10分）
            required_skills = shift.get("required_skills", [])
            emp_skills = emp.get("skills", [])
            if required_skills:
                if any(skill in emp_skills for skill in required_skills):
                    score -= 10
                else:
                    score += 50  # 没有技能的惩罚
            
            # 偏好班次：偏好该班次的优先（减少5分）
            preferred_shifts = emp.get("preferred_shifts", [])
            if shift_id in preferred_shifts:
                score -= 5
            
            # 工作小时平衡：工作小时少的优先
            score += current_workload[emp_id]["hours"] * 0.1
            
            # 【优化】休息周期内班次一致 + 休息后轮换
            cycle_info = employee_shift_cycle.get(emp_id, {})
            new_type = get_shift_type(shift_id)
            
            # 早早班员工固定，不参与轮换逻辑
            is_early_early_employee = emp_id in fixed_early_early_employees
            if is_early_early_employee:
                return score
            
            current_type = cycle_info.get("current_type")
            just_rested = check_if_rested(emp_id, date_str)
            
            if just_rested and current_type and current_type in ["早班", "晚班"]:
                # 刚休息过，优先分配相反的班次类型
                preferred_type = get_opposite_shift_type(current_type)
                if new_type == preferred_type:
                    score -= 5  # 轻微优先轮换后的班次
            elif current_type:
                # 同一休息周期内，优先保持相同班次类型
                if new_type == current_type:
                    score -= 3  # 轻微优先相同班次
            
            return score
        
        def get_required_staff(shift_id: str, date_str: str = None, schedule: dict = None) -> int:
            """获取班次所需人数（考虑特殊配置和动态规则）"""
            # 特殊规则：如果二期水吧-早早班没有排，那么二期水吧-早班需要2个人
            early_shift_id = "二期水吧-早班"
            early_early_shift_id = "二期水吧-早早班"
            
            if shift_id == early_shift_id and date_str and schedule:
                # 检查当天早早班是否有人
                early_early_count = schedule.get(date_str, {}).get("shift_counts", {}).get(early_early_shift_id, 0)
                if early_early_count == 0:
                    # 早早班没有人，早班需要2个人
                    return 2
                else:
                    # 早早班有人，早班只需要1个人
                    return 1
            
            # 其他班次使用默认配置
            if shift_id in flexible_shifts and "required_staff" in flexible_shifts[shift_id]:
                return flexible_shifts[shift_id]["required_staff"]
            return shifts[shift_id].get("required_staff", 1)
        
        def is_allow_empty(shift_id: str) -> bool:
            """检查班次是否允许为空"""
            if shift_id in flexible_shifts:
                return flexible_shifts[shift_id].get("allow_empty", False)
            return False
        
        def assign_employee_to_shift(date_str: str, shift_id: str, schedule: dict, current_workload: dict) -> bool:
            """尝试为班次分配一个员工，返回是否成功"""
            shift = shifts[shift_id]
            required_staff = get_required_staff(shift_id, date_str, schedule)  # 使用特殊配置，传入日期和排班表
            current_count = schedule[date_str]["shift_counts"].get(shift_id, 0)
            
            if current_count >= required_staff:
                return False  # 已经满了
            
            # 获取可用员工列表（排除已在该天有班次的）
            available_employees = [
                emp_id for emp_id in employee_list
                if emp_id not in schedule[date_str]["assignments"]
            ]
            
            # 计算每个员工的分数并排序
            candidates = []
            for emp_id in available_employees:
                conflicts = check_conflicts(emp_id, shift_id, date_str, schedule)
                if not conflicts:
                    score = get_employee_score(emp_id, shift_id, date_str, current_workload)
                    candidates.append((score, emp_id))
            
            # 按分数排序（分数低的优先）
            candidates.sort(key=lambda x: x[0])
            
            # 分配第一个候选员工
            if candidates:
                best_emp_id = candidates[0][1]
                schedule[date_str]["assignments"][best_emp_id] = shift_id
                schedule[date_str]["shift_counts"][shift_id] = current_count + 1
                
                # 更新工作量
                shift_duration = shift.get("duration_hours", 8)
                current_workload[best_emp_id]["days"] += 1
                current_workload[best_emp_id]["hours"] += shift_duration
                
                # 更新员工休息周期班次记录
                new_shift_type = get_shift_type(shift_id)
                cycle_info = employee_shift_cycle.get(best_emp_id, {})
                just_rested = check_if_rested(best_emp_id, date_str)
                
                if just_rested:
                    # 刚休息过，开始新的休息周期，更新班次类型
                    employee_shift_cycle[best_emp_id] = {
                        "current_type": new_shift_type,
                        "last_work_date": date_str,
                        "rested": True
                    }
                else:
                    # 同一休息周期内，保持班次类型，更新工作日期
                    if not cycle_info.get("current_type"):
                        cycle_info["current_type"] = new_shift_type
                    cycle_info["last_work_date"] = date_str
                    cycle_info["rested"] = False
                    employee_shift_cycle[best_emp_id] = cycle_info
                
                return True
            
            return False
        
        # 第一轮：优先处理固定早早班和特殊规则
        for date_str in schedule.keys():
            weekday_chinese = get_weekday_chinese(date_str)
            is_monday = weekday_chinese == "周一"
            
            # 特殊处理：周一不需要早早班
            if is_monday and monday_no_early_early:
                # 周一直接跳过早早班，不分配
                continue
            
            # 处理固定早早班
            if fixed_early_early_shift in shift_list:
                shift = shifts[fixed_early_early_shift]
                required_staff = shift.get("required_staff", 2)
                
                # 优先分配固定早早班人员
                for emp_id in fixed_early_early_employees:
                    if schedule[date_str]["shift_counts"].get(fixed_early_early_shift, 0) >= required_staff:
                        break
                    if emp_id not in schedule[date_str]["assignments"]:
                        conflicts = check_conflicts(emp_id, fixed_early_early_shift, date_str, schedule)
                        if not conflicts:
                            schedule[date_str]["assignments"][emp_id] = fixed_early_early_shift
                            schedule[date_str]["shift_counts"][fixed_early_early_shift] = \
                                schedule[date_str]["shift_counts"].get(fixed_early_early_shift, 0) + 1
                            shift_duration = shift.get("duration_hours", 8)
                            employee_workload[emp_id]["days"] += 1
                            employee_workload[emp_id]["hours"] += shift_duration
                            
                            # 更新员工休息周期记录（早早班员工固定，但仍需记录工作日期）
                            new_shift_type = get_shift_type(fixed_early_early_shift)
                            employee_shift_cycle[emp_id] = {
                                "current_type": new_shift_type,
                                "last_work_date": date_str,
                                "rested": False
                            }
        
        # 第二轮：按日期顺序分配所有班次（优先有技能的员工）
        for date_str in schedule.keys():
            weekday_chinese = get_weekday_chinese(date_str)
            is_monday = weekday_chinese == "周一"
            
            # 确定当天的班次列表（周一不需要早早班）
            if is_monday and monday_no_early_early:
                shift_list_today = [s for s in shift_list if s != fixed_early_early_shift]
            else:
                shift_list_today = shift_list
            
            # 对每个班次进行分配
            for shift_id in shift_list_today:
                if shift_id == fixed_early_early_shift:
                    # 早早班已在第一轮处理（非周一）
                    continue
                
                shift = shifts[shift_id]
                required_staff = get_required_staff(shift_id, date_str, schedule)  # 使用特殊配置，传入日期和排班表
                current_count = schedule[date_str]["shift_counts"].get(shift_id, 0)
                
                # 尝试填满该班次
                while current_count < required_staff:
                    if assign_employee_to_shift(date_str, shift_id, schedule, employee_workload):
                        current_count += 1
                    else:
                        break  # 无法再分配，跳出
        
        # 第三轮：填补空岗（放宽技能要求，优先填满岗位）
        max_iterations = 3  # 最多尝试3轮
        for iteration in range(max_iterations):
            vacancies_filled = 0
            
            for date_str in schedule.keys():
                weekday_chinese = get_weekday_chinese(date_str)
                is_monday = weekday_chinese == "周一"
                
                # 确定当天的班次列表（周一不需要早早班）
                if is_monday and monday_no_early_early:
                    shift_list_today = [s for s in shift_list if s != fixed_early_early_shift]
                else:
                    shift_list_today = shift_list
                
                for shift_id in shift_list_today:
                    shift = shifts[shift_id]
                    required_staff = get_required_staff(shift_id, date_str, schedule)  # 使用特殊配置，传入日期和排班表
                    current_count = schedule[date_str]["shift_counts"].get(shift_id, 0)
                    
                    if current_count < required_staff:
                        # 有空岗，尝试填补
                        if assign_employee_to_shift(date_str, shift_id, schedule, employee_workload):
                            vacancies_filled += 1
                        elif is_allow_empty(shift_id):
                            # 允许为空的班次，如果找不到人，跳过（不强制填满）
                            pass
            
            if vacancies_filled == 0:
                break  # 没有空岗需要填补，退出循环
        
        # 第四轮：为没有班的员工分配工作（平衡工作量）
        for date_str in schedule.keys():
            weekday_chinese = get_weekday_chinese(date_str)
            is_monday = weekday_chinese == "周一"
            
            # 确定当天的班次列表（周一不需要早早班）
            if is_monday and monday_no_early_early:
                shift_list_today = [s for s in shift_list if s != fixed_early_early_shift]
            else:
                shift_list_today = shift_list
            
            # 找出该天没有班的员工
            assigned_employees = set(schedule[date_str]["assignments"].keys())
            unassigned_employees = [emp_id for emp_id in employee_list if emp_id not in assigned_employees]
            
            # 按工作量排序（工作少的优先）
            unassigned_employees.sort(key=lambda e: (
                employee_workload[e]["days"],
                employee_workload[e]["hours"]
            ))
            
            # 尝试为这些员工找到合适的班次
            for emp_id in unassigned_employees:
                # 找出需要更多人的班次
                shift_shortages = []
                for shift_id in shift_list_today:
                    shift = shifts[shift_id]
                    required_staff = get_required_staff(shift_id, date_str, schedule)  # 使用特殊配置，传入日期和排班表
                    current_count = schedule[date_str]["shift_counts"].get(shift_id, 0)
                    
                    if current_count < required_staff:
                        shift_shortages.append((shift_id, required_staff - current_count))
                
                # 尝试分配（优先选择与当前休息周期班次类型一致的，或休息后轮换的）
                cycle_info = employee_shift_cycle.get(emp_id, {})
                current_type = cycle_info.get("current_type")
                just_rested = check_if_rested(emp_id, date_str)
                
                # 根据休息周期逻辑排序候选班次
                if current_type and current_type in ["早班", "晚班"]:
                    if just_rested:
                        # 刚休息过，优先轮换到相反班次
                        preferred = get_opposite_shift_type(current_type)
                        shift_shortages.sort(key=lambda x: (0 if get_shift_type(x[0]) == preferred else 1, -x[1]))
                    else:
                        # 同一休息周期内，优先相同班次
                        shift_shortages.sort(key=lambda x: (0 if get_shift_type(x[0]) == current_type else 1, -x[1]))
                
                for shift_id, shortage in shift_shortages:
                    conflicts = check_conflicts(emp_id, shift_id, date_str, schedule)
                    if not conflicts:
                        schedule[date_str]["assignments"][emp_id] = shift_id
                        schedule[date_str]["shift_counts"][shift_id] = \
                            schedule[date_str]["shift_counts"].get(shift_id, 0) + 1
                        shift = shifts[shift_id]
                        shift_duration = shift.get("duration_hours", 8)
                        employee_workload[emp_id]["days"] += 1
                        employee_workload[emp_id]["hours"] += shift_duration
                        
                        # 更新员工休息周期班次记录
                        new_shift_type = get_shift_type(shift_id)
                        cycle_info = employee_shift_cycle.get(emp_id, {})
                        just_rested = check_if_rested(emp_id, date_str)
                        
                        if just_rested:
                            employee_shift_cycle[emp_id] = {
                                "current_type": new_shift_type,
                                "last_work_date": date_str,
                                "rested": True
                            }
                        else:
                            if not cycle_info.get("current_type"):
                                cycle_info["current_type"] = new_shift_type
                            cycle_info["last_work_date"] = date_str
                            employee_shift_cycle[emp_id] = cycle_info
                        
                        break  # 分配成功，跳出循环
        
        # 保存排班表
        save_json(SCHEDULE_FILE, schedule)
        st.session_state.schedule = schedule
        st.success(f"✅ 排班表已成功生成！\n\n📅 日期范围：{start_date} 至 {end_date}（共 {len(date_range)} 天）\n\n现在可以在「查看排班」页面查看和导出排班表。")
        st.balloons()  # 庆祝动画
        st.rerun()


# HTML模板（Jinja2）用于生成海报级排班表
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>排班表</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
            padding: 40px;
            min-height: 100vh;
        }
        
        .poster-container {
            max-width: 100%;
            width: 100%;
            margin: 0 auto;
            background: #ffffff;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.15);
            overflow: visible;
        }
        
        .header {
            background: linear-gradient(135deg, #003366 0%, #004d99 100%);
            color: #ffffff;
            padding: 32px 40px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .header-left {
            display: flex;
            align-items: center;
            gap: 16px;
        }
        
        .header-icon {
            font-size: 36px;
        }
        
        .header-title {
            font-size: 28px;
            font-weight: 700;
            letter-spacing: 1px;
        }
        
        .header-subtitle {
            font-size: 14px;
            opacity: 0.9;
            margin-top: 4px;
        }
        
        .header-right {
            text-align: right;
        }
        
        .header-period {
            font-size: 20px;
            font-weight: 600;
            margin-bottom: 4px;
        }
        
        .header-dates {
            font-size: 14px;
            opacity: 0.85;
        }
        
        .schedule-grid {
            display: grid;
            grid-template-columns: 120px repeat({{ date_count }}, 1fr);
            gap: 0;
            background: #f8f9fa;
        }
        
        .grid-header {
            background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%);
            color: #ffffff;
            padding: 16px 12px;
            text-align: center;
            font-weight: 600;
            font-size: 13px;
            border-right: 1px solid rgba(255, 255, 255, 0.1);
            position: sticky;
            top: 0;
            z-index: 10;
        }
        
        .grid-header:first-child {
            border-left: none;
        }
        
        .grid-header:last-child {
            border-right: none;
        }
        
        .name-cell {
            background: linear-gradient(135deg, #ecf0f1 0%, #bdc3c7 100%);
            padding: 16px 12px;
            text-align: center;
            font-weight: 600;
            font-size: 14px;
            color: #2c3e50;
            border-right: 2px solid #95a5a6;
            border-bottom: 1px solid #bdc3c7;
            position: sticky;
            left: 0;
            z-index: 5;
        }
        
        .shift-cell {
            padding: 12px 8px;
            border-right: 1px solid #e0e0e0;
            border-bottom: 1px solid #e0e0e0;
            min-height: 60px;
            display: flex;
            align-items: center;
            justify-content: center;
            background: #ffffff;
            position: relative;
        }
        
        .shift-card {
            border-radius: 8px;
            padding: 10px 14px;
            font-size: 13px;
            font-weight: 600;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 6px;
            position: relative;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
            border-left: 4px solid;
            min-width: 80px;
            width: 100%;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }
        
        .shift-early {
            background: linear-gradient(135deg, #e0f7fa 0%, #80deea 100%);
            color: #00695c;
            border-left-color: #00897b;
        }
        
        .shift-late {
            background: linear-gradient(135deg, #e3f2fd 0%, #90caf9 100%);
            color: #1565c0;
            border-left-color: #1976d2;
        }
        
        .shift {
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            line-height: 1.4;
        }
        
        .fixed-rest {
            background: linear-gradient(135deg, #424242 0%, #616161 100%);
            color: #ffffff;
            border-left-color: #212121;
        }
        
        .no-role {
            background: #ffffff;
            color: #9e9e9e;
            border-left-color: #e0e0e0;
            border: 1px dashed #bdbdbd;
            box-shadow: none;
        }
        
        .shift-icon {
            font-size: 16px;
        }
        
        .footer {
            background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
            padding: 24px 40px;
            text-align: center;
            border-top: 1px solid #dee2e6;
        }
        
        .footer-text {
            font-size: 13px;
            color: #6c757d;
            margin-bottom: 8px;
        }
        
        .footer-slogan {
            font-size: 16px;
            color: #495057;
            font-weight: 600;
            letter-spacing: 1px;
        }
        
        .weekday-badge {
            font-size: 11px;
            opacity: 0.7;
            margin-top: 2px;
        }
    </style>
</head>
<body>
    <div class="poster-container">
        <div class="header">
            <div class="header-left">
                <span class="header-icon">🌊</span>
                <div>
                    <div class="header-title">沈阳水吧 | 智能排班公示</div>
                    <div class="header-subtitle">Intelligent Scheduling System</div>
                </div>
            </div>
            <div class="header-right">
                <div class="header-period">{{ period_title }}</div>
                <div class="header-dates">{{ date_range_text }}</div>
            </div>
        </div>
        
        <div class="schedule-grid">
            <!-- 表头 -->
            <div class="grid-header">姓名</div>
            {% for date_header in date_headers %}
            <div class="grid-header">
                {{ date_header.date }}<br>
                <span class="weekday-badge">{{ date_header.weekday }}</span>
            </div>
            {% endfor %}
            
            <!-- 数据行 -->
            {% for employee in employees %}
            <div class="name-cell">{{ employee.name }}</div>
            {% for shift_item in employee.shifts %}
            <div class="shift-cell">
                {% if shift_item.type == "shift" %}
                    {% if '早' in shift_item.value or '早早' in shift_item.value %}
                    <div class="shift-card shift-early shift">
                        <span class="shift-icon">☀️</span>
                        <span>{{ shift_item.value }}</span>
                    </div>
                    {% elif '晚' in shift_item.value %}
                    <div class="shift-card shift-late shift">
                        <span class="shift-icon">🌙</span>
                        <span>{{ shift_item.value }}</span>
                    </div>
                    {% else %}
                    <div class="shift-card shift-early shift">
                        <span>{{ shift_item.value }}</span>
                    </div>
                    {% endif %}
                {% elif shift_item.type == "fixed_rest" %}
                    <div class="shift-card fixed-rest">
                        <span class="shift-icon">🔒</span>
                        <span>固休</span>
                    </div>
                {% elif shift_item.type == "no_role" %}
                    <div class="shift-card no-role">
                        <span class="shift-icon">⚪️</span>
                        <span>无岗</span>
                    </div>
                {% endif %}
            </div>
            {% endfor %}
            {% endfor %}
        </div>
        
        <div class="footer">
            <div class="footer-text">生成时间：{{ generate_time }}</div>
            <div class="footer-slogan">高效协同，快乐工作 ✨</div>
        </div>
    </div>
</body>
</html>
"""


def generate_poster_image(schedule, employees, shifts, start_date_str, end_date_str, selected_employees=None):
    """生成海报级排班表图片
    
    Args:
        schedule: 排班数据
        employees: 员工数据
        shifts: 班次数据
        start_date_str: 开始日期
        end_date_str: 结束日期
        selected_employees: 可选，要显示的员工ID列表，None表示显示全部
    """
    if not HAS_JINJA2 or not HAS_PLAYWRIGHT:
        return None
    
    try:
        # 步骤A：数据清洗与准备
        dates = sorted(schedule.keys())
        
        # 构建日期表头
        date_headers = []
        for date_str in dates:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            weekday_map = {
                0: "周一", 1: "周二", 2: "周三", 3: "周四",
                4: "周五", 5: "周六", 6: "周日"
            }
            weekday = weekday_map[date_obj.weekday()]
            date_headers.append({
                "date": date_str[5:],  # 只显示月-日
                "weekday": weekday
            })
        
        # 构建员工和班次数据（支持员工筛选）
        employees_data = []
        for emp_id, emp in employees.items():
            # 如果指定了员工筛选，只处理选中的员工
            if selected_employees is not None and emp_id not in selected_employees:
                continue
                
            emp_name = emp.get("name", emp_id)
            shifts_list = []
            
            rest_day = emp.get("rest_day", "")  # 获取员工的固定休息日
            for date_str in dates:
                date_schedule = schedule.get(date_str, {})
                assignments = date_schedule.get("assignments", {})
                shift_id = assignments.get(emp_id, None)
                
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                weekday_map = {
                    0: "周一", 1: "周二", 2: "周三", 3: "周四",
                    4: "周五", 5: "周六", 6: "周日"
                }
                weekday_chinese = weekday_map[date_obj.weekday()]
                
                if shift_id:
                    shifts_list.append({"type": "shift", "value": shift_id})
                elif rest_day and weekday_chinese == rest_day:
                    shifts_list.append({"type": "fixed_rest", "value": None})
                else:
                    shifts_list.append({"type": "no_role", "value": None})
            
            employees_data.append({
                "name": emp_name,
                "shifts": shifts_list
            })
        
        # 计算周期标题
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        try:
            week_num = start_date.isocalendar()[1]
            period_title = f"{start_date.year}年{start_date.month}月 第{week_num}周"
        except:
            period_title = f"{start_date_str} 至 {end_date_str}"
        
        date_range_text = f"{start_date_str} ~ {end_date_str}"
        generate_time = datetime.now().strftime("%Y年%m月%d日 %H:%M")
        
        # 步骤B：Jinja2模板渲染
        template = Template(HTML_TEMPLATE)
        html_content = template.render(
            date_count=len(dates),
            date_headers=date_headers,
            employees=employees_data,
            period_title=period_title,
            date_range_text=date_range_text,
            generate_time=generate_time
        )
        
        # 步骤C：Playwright无头浏览器截图
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            
            # 设置较大的视口以保证清晰度
            # 设置较大的视口以保证清晰度（升级到1800px宽度）
            page.set_viewport_size({"width": 1800, "height": 2400})
            
            # 加载HTML内容
            page.set_content(html_content, wait_until="networkidle")
            
            # 等待页面渲染完成
            page.wait_for_timeout(1000)
            
            # 获取页面实际高度
            page_height = page.evaluate("() => Math.max(document.body.scrollHeight, document.body.offsetHeight, document.documentElement.clientHeight, document.documentElement.scrollHeight, document.documentElement.offsetHeight)")
            
            # 设置更大的视口以适应内容
            if page_height > 2400:
                page.set_viewport_size({"width": 1800, "height": int(page_height + 200)})
                page.wait_for_timeout(500)
            
            # 全页面截图
            screenshot_bytes = page.screenshot(full_page=True, type="png")
            
            browser.close()
        
        return screenshot_bytes
    
    except Exception as e:
        st.error(f"生成海报时出错：{str(e)}")
        return None


def export_schedule(format_type: str = "excel"):
    """导出排班表（简化版）"""
    schedule = st.session_state.schedule
    employees = st.session_state.employees
    shifts = st.session_state.shifts
    
    if not schedule:
        return None, "暂无排班数据"
    
    dates = sorted(schedule.keys())
    
    # 构建简洁的排班表：每行一个员工，每列一个日期
    emp_schedule = {}
    
    # 先收集每个员工在排班期间的班次类型（用于确定待岗类型）
    emp_shift_types = {}
    for emp_id, emp in employees.items():
        emp_shift_types[emp_id] = None
        # 找到该员工最近一次有班的班次类型
    for date_str in dates:
            assignments = schedule.get(date_str, {}).get("assignments", {})
            if emp_id in assignments:
                shift_id = assignments[emp_id]
                if "早早" in shift_id:
                    emp_shift_types[emp_id] = "早早班"
                elif "早" in shift_id:
                    emp_shift_types[emp_id] = "早班"
                elif "晚" in shift_id:
                    emp_shift_types[emp_id] = "晚班"
                break  # 找到第一个班次就确定类型
    
    for emp_id, emp in employees.items():
        emp_name = emp.get("name", emp_id)
        rest_day = emp.get("rest_day", "")
        emp_schedule[emp_name] = {}
        
        # 追踪当前工作周期的班次类型
        current_shift_type = emp_shift_types.get(emp_id, "早班")  # 默认早班
        
        for date_str in dates:
            date_schedule = schedule.get(date_str, {})
            assignments = date_schedule.get("assignments", {})
            weekday_chinese = get_weekday_chinese(date_str)
            
            if emp_id in assignments:
                # 有班次，显示班次名称（简化显示）
                shift_id = assignments[emp_id]
                # 简化班次名称显示
                if "早早" in shift_id:
                    emp_schedule[emp_name][date_str] = "早早班"
                    current_shift_type = "早早班"
                elif "早" in shift_id:
                    emp_schedule[emp_name][date_str] = "早班"
                    current_shift_type = "早班"
                elif "晚" in shift_id:
                    emp_schedule[emp_name][date_str] = "晚班"
                    current_shift_type = "晚班"
                else:
                    emp_schedule[emp_name][date_str] = shift_id
            elif rest_day and weekday_chinese == rest_day:
                emp_schedule[emp_name][date_str] = "休"
            else:
                # 待岗：根据当前工作周期的班次类型来标注
                if current_shift_type == "早早班":
                    emp_schedule[emp_name][date_str] = "待岗(早早)"
                elif current_shift_type == "早班":
                    emp_schedule[emp_name][date_str] = "待岗(早)"
                elif current_shift_type == "晚班":
                    emp_schedule[emp_name][date_str] = "待岗(晚)"
                else:
                    emp_schedule[emp_name][date_str] = "待岗"
    
    # 创建 DataFrame
    df = pd.DataFrame(emp_schedule).T  # 转置：员工为行，日期为列
    
    # 添加星期行作为表头
    weekday_row = {date_str: get_weekday_chinese(date_str) for date_str in dates}
    df_with_weekday = pd.DataFrame([weekday_row], index=["星期"])
    df = pd.concat([df_with_weekday, df])
    
    if format_type == "csv":
        csv_data = df.to_csv(encoding='utf-8-sig')
        filename = f"排班表_{dates[0]}_{dates[-1]}.csv"
        return csv_data.encode('utf-8-sig'), filename
    
    elif format_type == "excel":
        from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 为每个员工添加汇总统计列
        summary_data = {}
        for emp_name, emp_dates in emp_schedule.items():
            work_days = 0
            morning_count = 0
            early_early_count = 0
            evening_count = 0
            rest_days = 0
            standby_days = 0
            
            for date_str, shift_val in emp_dates.items():
                if "早早" in shift_val:
                    work_days += 1
                    early_early_count += 1
                elif "早班" in shift_val:
                    work_days += 1
                    morning_count += 1
                elif "晚班" in shift_val:
                    work_days += 1
                    evening_count += 1
                elif shift_val == "休":
                    rest_days += 1
                elif "待岗" in shift_val:
                    standby_days += 1
            
            summary_data[emp_name] = {
                "上班": work_days,
                "早班": morning_count,
                "早早": early_early_count,
                "晚班": evening_count,
                "休息": rest_days,
                "待岗": standby_days
            }
        
        # 添加汇总列到 DataFrame
        for col_name in ["上班", "早班", "早早", "晚班", "休息", "待岗"]:
            df[col_name] = ""
            df.loc["星期", col_name] = "汇总"
            for emp_name in emp_schedule.keys():
                if emp_name in df.index:
                    df.loc[emp_name, col_name] = summary_data.get(emp_name, {}).get(col_name, 0)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 写入主表，从第3行开始（留出标题行）
            df.to_excel(writer, sheet_name='排班表', startrow=2)
            
            worksheet = writer.sheets['排班表']
            
            # 添加专业标题
            title_font = Font(color="1F4E79", bold=True, size=16)
            subtitle_font = Font(color="5B9BD5", size=11)
            
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(10, len(dates) + 1))
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = f"📅 员工排班表"
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(10, len(dates) + 1))
            subtitle_cell = worksheet.cell(row=2, column=1)
            subtitle_cell.value = f"排班周期：{dates[0]} 至 {dates[-1]}  |  员工数量：{len(employees)}人  |  生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
            subtitle_cell.font = subtitle_font
            subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 定义样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            weekday_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            name_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            summary_header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            summary_fill = PatternFill(start_color="F2E7FE", end_color="F2E7FE", fill_type="solid")
            summary_font = Font(color="7030A0", bold=True)
            
            rest_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            rest_font = Font(color="006100", bold=True)
            
            standby_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            standby_font = Font(color="9C5700")
            
            morning_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            morning_font = Font(color="1F4E79", bold=True)
            
            evening_fill = PatternFill(start_color="E2D5F1", end_color="E2D5F1", fill_type="solid")
            evening_font = Font(color="5B2C6F", bold=True)
            
            early_early_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            early_early_font = Font(color="C55A11", bold=True)
            
            thin_border = Border(
                left=Side(style='thin', color='B4B4B4'), 
                right=Side(style='thin', color='B4B4B4'), 
                top=Side(style='thin', color='B4B4B4'), 
                bottom=Side(style='thin', color='B4B4B4')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 计算汇总列的起始位置
            summary_start_col = len(dates) + 2  # +1 for index col, +1 for 1-based

            # 遍历单元格应用样式（从第3行开始，因为前2行是标题）
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    
                    val = str(cell.value) if cell.value else ""
                    
                    # 判断是否是汇总列
                    is_summary_col = cell.column >= summary_start_col
                    
                    # 第3行：表头（日期行）
                    if cell.row == 3:
                        if is_summary_col:
                            cell.fill = summary_header_fill
                        else:
                            cell.fill = header_fill
                        cell.font = header_font
                    # 第4行：星期行
                    elif cell.row == 4:
                        if is_summary_col:
                            cell.fill = summary_header_fill
                        else:
                            cell.fill = weekday_fill
                        cell.font = header_font
                    # 第一列：员工姓名
                    elif cell.column == 1:
                        cell.fill = name_fill
                        cell.font = header_font
                    # 汇总列
                    elif is_summary_col and cell.row > 4:
                        cell.fill = summary_fill
                        cell.font = summary_font
                    # 内容单元格根据班次类型着色
                    elif "休" in val:
                        cell.fill = rest_fill
                        cell.font = rest_font
                    elif "待岗" in val:
                        cell.fill = standby_fill
                        cell.font = standby_font
                    elif "早早" in val:
                        cell.fill = early_early_fill
                        cell.font = early_early_font
                    elif "早" in val:
                        cell.fill = morning_fill
                        cell.font = morning_font
                    elif "晚" in val:
                        cell.fill = evening_fill
                        cell.font = evening_font

            # 调整列宽和行高
            worksheet.column_dimensions['A'].width = 10
            for col_idx in range(2, worksheet.max_column + 1):
                col_letter = get_column_letter(col_idx)
                if col_idx >= summary_start_col:
                    worksheet.column_dimensions[col_letter].width = 8
                else:
                    worksheet.column_dimensions[col_letter].width = 12
            
            # 设置行高
            worksheet.row_dimensions[1].height = 28  # 标题行
            worksheet.row_dimensions[2].height = 20  # 副标题行
            for row_idx in range(3, worksheet.max_row + 1):
                worksheet.row_dimensions[row_idx].height = 22
            
            # 冻结窗格：固定第一列和前4行（包括新增的标题行）
            worksheet.freeze_panes = 'B5'
            
            # 添加图例说明（在数据下方）
            legend_row = worksheet.max_row + 2
            legend_items = [
                ("图例说明：", None, Font(bold=True, size=11)),
                ("早班", morning_fill, morning_font),
                ("早早班", early_early_fill, early_early_font),
                ("晚班", evening_fill, evening_font),
                ("休息", rest_fill, rest_font),
                ("待岗", standby_fill, standby_font),
            ]
            
            for i, (text, fill, font) in enumerate(legend_items):
                cell = worksheet.cell(row=legend_row, column=i + 1)
                cell.value = text
                cell.font = font
                if fill:
                    cell.fill = fill
                cell.alignment = center_alignment
                cell.border = thin_border
        
        excel_data = output.getvalue()
        filename = f"排班表_{dates[0]}_{dates[-1]}.xlsx"
        return excel_data, filename
    
    return None, "不支持的格式"



def ai_schedule_tuning():
    """AI 智能微调排班表"""
    st.header("🤖 AI 智能微调")
    
    schedule = st.session_state.schedule
    employees = st.session_state.employees
    shifts = st.session_state.shifts
    
    if not schedule:
        st.info("💡 暂无排班表，请先在「生成排班」页面生成排班表")
        return
    
    if not HAS_OPENAI:
        st.warning("⚠️ 需要安装 openai 库才能使用 AI 微调功能")
        st.code("pip install openai", language="bash")
        return
    
    api_key = st.session_state.get("ai_api_key", "")
    base_url = st.session_state.get("ai_base_url", "https://generativelanguage.googleapis.com/v1beta/openai/")
    proxy = st.session_state.get("ai_proxy", "")
    
    if not api_key:
        st.warning("⚠️ 请先在左侧边栏展开「🤖 AI 配置」，填写您的 API Key")
        st.info("💡 **获取 API Key 的方法：**\n1. 访问 [Google AI Studio](https://aistudio.google.com/apikey)\n2. 创建或复制您的 API Key\n3. 粘贴到左侧的配置框中")
        return
    
    # 准备当前排班数据
    current_schedule_data = {
        "schedule": schedule,
        "employees": employees,
        "shifts": shifts
    }
    
    # 显示当前排班概况
    dates = sorted(schedule.keys())
    st.info(f"📅 当前排班范围：{dates[0]} 至 {dates[-1]}（共 {len(dates)} 天，{len(employees)} 名员工）")
    
    st.subheader("📝 输入修改指令")
    
    # 显示员工列表供参考
    with st.expander("👥 查看员工列表（点击展开）"):
        emp_names = [f"• {emp.get('name', emp_id)}" for emp_id, emp in employees.items()]
        st.markdown("  \n".join(emp_names))
    
    user_instruction = st.text_area(
        "请用自然语言描述您想要的排班修改",
        placeholder="例如：\n- 让张三在1月15日上早班\n- 把李四1月20日的班次改为晚班\n- 将王五1月10日的班次取消\n- 交换张三和李四在1月18日的班次",
        height=150,
        help="用自然语言描述您想要的排班修改，AI 会分析并生成新的排班表"
    )
    
    if st.button("🚀 执行 AI 微调", type="primary", use_container_width=True):
        if not user_instruction.strip():
            st.error("❌ 请输入修改指令")
            return
        
        with st.spinner("🤖 AI 正在分析和修改排班表，请耐心等待..."):
            try:
                # 设置代理（如果用户配置了）
                original_proxy = os.environ.get("HTTP_PROXY", None)
                if proxy:
                    os.environ["HTTP_PROXY"] = proxy
                    os.environ["HTTPS_PROXY"] = proxy
                
                try:
                    # 构建 Prompt
                    prompt = f"""你是一个专业的排班系统助手。请根据用户的指令修改排班表。

当前排班数据（JSON格式）：
{json.dumps(current_schedule_data, ensure_ascii=False, indent=2)}

用户指令：
{user_instruction}

要求：
1. 仔细分析用户的指令，理解用户想要修改的内容
2. 仅修改用户指定的部分，保持其他排班不变
3. 确保修改后的排班表符合业务规则（员工技能匹配、休息日等）
4. 返回完整的修改后的 JSON 数据，格式与输入完全一致
5. 只返回 JSON 数据，不要添加任何解释性文字

请返回修改后的 JSON："""
                    
                    # 调用 OpenAI API
                    client = OpenAI(
                        api_key=api_key,
                        base_url=base_url
                    )
                    
                    # 获取用户配置的模型名称
                    model_name = st.session_state.get("ai_model", "gemini-2.0-flash")
                    
                    response = client.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "system", "content": "你是一个专业的排班系统助手，擅长理解和修改排班表数据。"},
                            {"role": "user", "content": prompt}
                        ],
                        temperature=0.3
                    )
                    
                    ai_response_text = response.choices[0].message.content.strip()
                    
                    # 尝试提取 JSON（去除可能的代码块标记）
                    if "```json" in ai_response_text:
                        ai_response_text = ai_response_text.split("```json")[1].split("```")[0].strip()
                    elif "```" in ai_response_text:
                        ai_response_text = ai_response_text.split("```")[1].split("```")[0].strip()
                    
                    # 解析 AI 返回的 JSON
                    new_schedule_data = json.loads(ai_response_text)
                    
                    # 保存到 session_state 用于对比
                    st.session_state.ai_modified_schedule = new_schedule_data.get("schedule", {})
                    st.session_state.ai_instruction = user_instruction
                    
                    st.success("✅ AI 分析完成！")
                    
                finally:
                    # 恢复原始代理设置
                    if proxy:
                        if original_proxy:
                            os.environ["HTTP_PROXY"] = original_proxy
                            os.environ["HTTPS_PROXY"] = original_proxy
                        else:
                            os.environ.pop("HTTP_PROXY", None)
                            os.environ.pop("HTTPS_PROXY", None)
                
                st.rerun()
                
            except json.JSONDecodeError as e:
                st.error(f"❌ AI 返回的数据格式错误：{str(e)}")
                if "ai_response_text" in locals():
                    st.code(ai_response_text, language="text")
            except Exception as e:
                st.error(f"❌ AI 微调失败：{str(e)}")
                import traceback
                st.code(traceback.format_exc(), language="python")
    
    # 显示 Diff 对比视图
    if "ai_modified_schedule" in st.session_state and st.session_state.ai_modified_schedule:
        st.markdown("---")
        st.subheader("📊 修改对比")
        
        modified_schedule = st.session_state.ai_modified_schedule
        instruction = st.session_state.get("ai_instruction", "")
        
        st.info(f"**修改指令**：{instruction}")
        
        # 计算差异
        dates = sorted(set(list(schedule.keys()) + list(modified_schedule.keys())))
        diff_data = []
        
        for date_str in dates:
            original_assignments = schedule.get(date_str, {}).get("assignments", {})
            modified_assignments = modified_schedule.get(date_str, {}).get("assignments", {})
            
            all_emp_ids = set(list(original_assignments.keys()) + list(modified_assignments.keys()))
            
            for emp_id in all_emp_ids:
                original_shift = original_assignments.get(emp_id)
                modified_shift = modified_assignments.get(emp_id)
                emp = employees.get(emp_id, {})
                emp_name = emp.get("name", emp_id)
                
                if original_shift != modified_shift:
                    diff_data.append({
                        "日期": date_str,
                        "员工": emp_name,
                        "修改前": original_shift if original_shift else "无",
                        "修改后": modified_shift if modified_shift else "无",
                        "状态": "🔄 已修改" if (original_shift and modified_shift) else ("➕ 新增" if not original_shift else "➖ 删除")
                    })
        
        if diff_data:
            diff_df = pd.DataFrame(diff_data)
            st.dataframe(diff_df, use_container_width=True, hide_index=True)
            
            col_confirm, col_cancel = st.columns([1, 1])
            with col_confirm:
                if st.button("✅ 确认并应用修改", type="primary", use_container_width=True):
                    # 保存修改后的排班表
                    save_json(SCHEDULE_FILE, modified_schedule)
                    st.session_state.schedule = modified_schedule
                    # 清除临时数据
                    if "ai_modified_schedule" in st.session_state:
                        del st.session_state.ai_modified_schedule
                    if "ai_instruction" in st.session_state:
                        del st.session_state.ai_instruction
                    st.success("✅ 排班表已更新！")
                    st.rerun()
            
            with col_cancel:
                if st.button("❌ 取消修改", use_container_width=True):
                    if "ai_modified_schedule" in st.session_state:
                        del st.session_state.ai_modified_schedule
                    if "ai_instruction" in st.session_state:
                        del st.session_state.ai_instruction
                    st.rerun()
        else:
            st.info("💡 AI 未检测到需要修改的内容，排班表保持不变")



def get_required_staff_for_view(shift_id: str, date_str: str, schedule: dict, shifts: dict) -> int:
    """获取班次所需人数（考虑动态规则，用于查看排班）"""
    # 特殊规则：如果二期水吧-早早班没有排，那么二期水吧-早班需要2个人
    early_shift_id = "二期水吧-早班"
    early_early_shift_id = "二期水吧-早早班"
    
    if shift_id == early_shift_id:
        # 检查当天早早班是否有人
        early_early_count = schedule.get(date_str, {}).get("shift_counts", {}).get(early_early_shift_id, 0)
        if early_early_count == 0:
            # 早早班没有人，早班需要2个人
            return 2
        else:
            # 早早班有人，早班只需要1个人
            return 1
    
    # 其他班次使用默认配置
    return shifts[shift_id].get("required_staff", 1)


def view_schedule():
    """查看排班表"""
    st.header("📋 查看排班表")
    
    schedule = st.session_state.schedule
    employees = st.session_state.employees
    shifts = st.session_state.shifts
    
    if not schedule:
        st.markdown("""
        <div style="text-align: center; padding: 60px 20px; background: #f8f9fa; border-radius: 12px;">
            <h2 style="color: #6c757d;">📭 暂无排班表</h2>
            <p style="color: #adb5bd;">请先在「🚀 生成排班」页面生成排班表</p>
        </div>
        """, unsafe_allow_html=True)
        return
    
    dates = sorted(schedule.keys())
    
    # 计算统计数据
    total_assignments = sum(len(d.get("assignments", {})) for d in schedule.values())
    total_vacancies = 0
    for date_str in dates:
        weekday_chinese = get_weekday_chinese(date_str)
        for shift_id in shifts:
            required = get_required_staff_for_view(shift_id, date_str, schedule, shifts)
            actual = sum(1 for s in schedule[date_str].get("assignments", {}).values() if s == shift_id)
            
            # 计算该班次在固定休息日有多少可用员工（这些员工不算缺人）
            # 只统计有相应技能且当天是固定休息日的员工
            rest_day_available = 0
            shift = shifts.get(shift_id, {})
            required_skills = shift.get("required_skills", [])
            
            for emp_id, emp in employees.items():
                if emp.get("rest_day") == weekday_chinese:
                    # 检查这个员工是否有这个班次所需的技能
                    emp_skills = emp.get("skills", [])
                    if not required_skills or any(skill in emp_skills for skill in required_skills):
                        rest_day_available += 1
            
            # 空岗 = 需要人数 - 实际上班人数 - 固定休息日可用人数
            # 如果固定休息日的人数能填补缺口，不算空岗
            shortage = required - actual
            if shortage > 0:
                # 扣除固定休息日可用人数后，才是真正的空岗
                real_shortage = max(0, shortage - rest_day_available)
                total_vacancies += real_shortage
    
    # 顶部统计卡片（响应式）
    st.markdown(f"""
    <div style="display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 20px;">
        <div class="stat-card" style="flex: 1 1 150px; min-width: 120px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    padding: 18px 12px; border-radius: 12px; color: white; text-align: center; box-shadow: 0 4px 15px rgba(102,126,234,0.3);">
            <h2 style="margin: 0; color: white; font-size: 28px;">{len(dates)}</h2>
            <p style="margin: 6px 0 0 0; opacity: 0.9; font-size: 13px;">排班天数</p>
        </div>
        <div class="stat-card" style="flex: 1 1 150px; min-width: 120px; background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); 
                    padding: 18px 12px; border-radius: 12px; color: white; text-align: center; box-shadow: 0 4px 15px rgba(17,153,142,0.3);">
            <h2 style="margin: 0; color: white; font-size: 28px;">{total_assignments}</h2>
            <p style="margin: 6px 0 0 0; opacity: 0.9; font-size: 13px;">总排班次数</p>
        </div>
        <div class="stat-card" style="flex: 1 1 150px; min-width: 120px; background: linear-gradient(135deg, {'#f093fb' if total_vacancies == 0 else '#eb3349'} 0%, {'#f5576c' if total_vacancies == 0 else '#f45c43'} 100%); 
                    padding: 18px 12px; border-radius: 12px; color: white; text-align: center; box-shadow: 0 4px 15px rgba(245,87,108,0.3);">
            <h2 style="margin: 0; color: white; font-size: 28px;">{total_vacancies}</h2>
            <p style="margin: 6px 0 0 0; opacity: 0.9; font-size: 13px;">{'✅ 无空岗' if total_vacancies == 0 else '⚠️ 空岗数'}</p>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # 导出按钮
    col1, col2, col3 = st.columns([2, 1, 1])
    with col2:
        excel_data, excel_filename = export_schedule("excel")
        if excel_data:
            st.download_button("📊 导出 Excel", excel_data, excel_filename, 
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             use_container_width=True, type="primary")
    with col3:
        csv_data, csv_filename = export_schedule("csv")
        if csv_data:
            st.download_button("📄 导出 CSV", csv_data, csv_filename, "text/csv", use_container_width=True)
    
    st.markdown("---")
    
    # 员工个人排班查询（新功能）
    st.subheader("🔍 查询排班")
    emp_names = ["全部员工"] + [emp.get("name", emp_id) for emp_id, emp in employees.items()]
    selected_emp = st.selectbox("选择员工", emp_names, label_visibility="collapsed")
    
    if selected_emp != "全部员工":
        # 显示单个员工的排班
        emp_id = None
        for eid, emp in employees.items():
            if emp.get("name") == selected_emp:
                emp_id = eid
                break
        
        if emp_id:
            emp = employees[emp_id]
            rest_day = emp.get("rest_day", "无")
            emp_shifts = []
            
            # 先确定该员工当前工作周期的班次类型
            current_shift_type = "早班"  # 默认
            for date_str in dates:
                assignments = schedule.get(date_str, {}).get("assignments", {})
                if emp_id in assignments:
                    shift_id = assignments[emp_id]
                    if "早早" in shift_id:
                        current_shift_type = "早早班"
                    elif "早" in shift_id:
                        current_shift_type = "早班"
                    elif "晚" in shift_id:
                        current_shift_type = "晚班"
                    break
            
            for date_str in dates:
                assignments = schedule.get(date_str, {}).get("assignments", {})
                weekday = get_weekday_chinese(date_str)
                if emp_id in assignments:
                    shift_id = assignments[emp_id]
                    # 更新当前班次类型
                    if "早早" in shift_id:
                        current_shift_type = "早早班"
                        display_shift = "早早班"
                    elif "早" in shift_id:
                        current_shift_type = "早班"
                        display_shift = "早班"
                    elif "晚" in shift_id:
                        current_shift_type = "晚班"
                        display_shift = "晚班"
                    else:
                        display_shift = shift_id
                    emp_shifts.append({"日期": date_str, "星期": weekday, "班次": display_shift})
                elif rest_day == weekday:
                    emp_shifts.append({"日期": date_str, "星期": weekday, "班次": "🔒 固休"})
                else:
                    # 待岗区分早/晚班
                    if current_shift_type == "早早班":
                        standby_text = "📍 待岗(早早)"
                    elif current_shift_type == "早班":
                        standby_text = "📍 待岗(早)"
                    elif current_shift_type == "晚班":
                        standby_text = "📍 待岗(晚)"
                    else:
                        standby_text = "📍 待岗"
                    emp_shifts.append({"日期": date_str, "星期": weekday, "班次": standby_text})
            
            st.markdown(f"**{selected_emp}** 的排班（固定休息日：{rest_day}）")
            emp_df = pd.DataFrame(emp_shifts)
            st.dataframe(emp_df, use_container_width=True, hide_index=True, height=300)
    else:
        # 添加图例说明（响应式）
        st.markdown("""
        <div style="background: #f8fafc; padding: 10px 14px; border-radius: 10px; margin-bottom: 15px;">
            <div style="display: flex; flex-wrap: wrap; align-items: center; gap: 8px;">
                <span style="font-size: 12px; color: #64748b; margin-right: 5px;">📋 图例：</span>
                <span class="legend-item" style="display: inline-flex; align-items: center; font-size: 12px;">
                    <span class="legend-color" style="width: 14px; height: 14px; border-radius: 3px; margin-right: 4px; background: linear-gradient(135deg, #DDEBF7 0%, #B8D4F0 100%);"></span>早班
                </span>
                <span class="legend-item" style="display: inline-flex; align-items: center; font-size: 12px;">
                    <span class="legend-color" style="width: 14px; height: 14px; border-radius: 3px; margin-right: 4px; background: linear-gradient(135deg, #FCE4D6 0%, #F8CBAD 100%);"></span>早早班
                </span>
                <span class="legend-item" style="display: inline-flex; align-items: center; font-size: 12px;">
                    <span class="legend-color" style="width: 14px; height: 14px; border-radius: 3px; margin-right: 4px; background: linear-gradient(135deg, #E2D5F1 0%, #D4C4E8 100%);"></span>晚班
                </span>
                <span class="legend-item" style="display: inline-flex; align-items: center; font-size: 12px;">
                    <span class="legend-color" style="width: 14px; height: 14px; border-radius: 3px; margin-right: 4px; background: linear-gradient(135deg, #C6EFCE 0%, #A9E4B1 100%);"></span>休息
                </span>
                <span class="legend-item" style="display: inline-flex; align-items: center; font-size: 12px;">
                    <span class="legend-color" style="width: 14px; height: 14px; border-radius: 3px; margin-right: 4px; background: linear-gradient(135deg, #FFEB9C 0%, #FFD966 100%);"></span>待岗
                </span>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # 显示所有员工的日历视图（带着色）
        calendar_data = []
        
        # 收集每个员工的班次类型用于待岗显示
        emp_shift_types = {}
        for emp_id, emp in employees.items():
            emp_shift_types[emp_id] = "早班"
            for date_str in dates:
                assignments = schedule.get(date_str, {}).get("assignments", {})
                if emp_id in assignments:
                    shift_id = assignments[emp_id]
                    if "早早" in shift_id:
                        emp_shift_types[emp_id] = "早早班"
                    elif "早" in shift_id:
                        emp_shift_types[emp_id] = "早班"
                    elif "晚" in shift_id:
                        emp_shift_types[emp_id] = "晚班"
                    break
        
        for date_str in dates:
            date_schedule = schedule[date_str]
            weekday_chinese = get_weekday_chinese(date_str)
            
            for emp_id, emp in employees.items():
                emp_name = emp.get("name", emp_id)
                rest_day = emp.get("rest_day", "")
                assignments = date_schedule.get("assignments", {})
                
                if emp_id in assignments:
                    shift_id = assignments[emp_id]
                    # 保留完整班次名称（包含岗位信息）
                    display_shift = shift_id
                    
                    # 记录类型用于待岗标记
                    if "早早" in shift_id:
                        emp_shift_types[emp_id] = "早早班"
                    elif "早" in shift_id:
                        emp_shift_types[emp_id] = "早班"
                    elif "晚" in shift_id:
                        emp_shift_types[emp_id] = "晚班"
                elif rest_day and weekday_chinese == rest_day:
                    display_shift = "休"
                else:
                    # 待岗显示
                    shift_type = emp_shift_types.get(emp_id, "早班")
                    if shift_type == "早早班":
                        display_shift = "待岗(早早)"
                    elif shift_type == "早班":
                        display_shift = "待岗(早)"
                    elif shift_type == "晚班":
                        display_shift = "待岗(晚)"
                    else:
                        display_shift = "待岗"
                
                calendar_data.append({
                    "日期": f"{date_str[5:]}({weekday_chinese[:1]})",
                    "员工": emp_name,
                    "班次": display_shift
                })
        
        if calendar_data:
            calendar_df = pd.DataFrame(calendar_data)
            pivot_table = calendar_df.pivot_table(
                index="员工", columns="日期", values="班次",
                aggfunc=lambda x: x.iloc[0] if len(x) > 0 else "", fill_value="—"
            )
            
            # 定义着色函数
            def style_shift_cell(val):
                """根据班次类型返回单元格样式"""
                val_str = str(val)
                if "早早" in val_str:
                    return 'background: linear-gradient(135deg, #FCE4D6 0%, #F8CBAD 100%); color: #C55A11; font-weight: 600;'
                elif "早" in val_str and "早早" not in val_str:
                    return 'background: linear-gradient(135deg, #DDEBF7 0%, #B8D4F0 100%); color: #1F4E79; font-weight: 600;'
                elif "晚" in val_str:
                    return 'background: linear-gradient(135deg, #E2D5F1 0%, #D4C4E8 100%); color: #5B2C6F; font-weight: 600;'
                elif val_str == "休":
                    return 'background: linear-gradient(135deg, #C6EFCE 0%, #A9E4B1 100%); color: #006100; font-weight: 600;'
                elif "待" in val_str:
                    return 'background: linear-gradient(135deg, #FFEB9C 0%, #FFD966 100%); color: #9C5700; font-weight: 500;'
                elif val_str == "—":
                    return 'background: #f8f9fa; color: #adb5bd;'
                else:
                    return ''
            
            # 应用样式
            styled_table = pivot_table.style.applymap(style_shift_cell)
            styled_table = styled_table.set_properties(**{
                'text-align': 'center',
                'font-size': '12px',
                'padding': '10px 4px',
                'border': '1px solid #e2e8f0',
                'min-width': '100px',
                'line-height': '1.2'
            })
            styled_table = styled_table.set_table_styles([
                {'selector': 'th', 'props': [
                    ('background', 'linear-gradient(135deg, #4472C4 0%, #5B9BD5 100%)'),
                    ('color', 'white'),
                    ('font-weight', '600'),
                    ('text-align', 'center'),
                    ('padding', '12px 6px'),
                    ('font-size', '12px'),
                    ('border', '1px solid #3a63ad')
                ]},
                {'selector': 'th.row_heading', 'props': [
                    ('background', 'linear-gradient(135deg, #2F5496 0%, #4472C4 100%)'),
                    ('color', 'white'),
                    ('font-weight', '600'),
                    ('min-width', '80px'),
                    ('position', 'sticky'),
                    ('left', '0'),
                    ('z-index', '1')
                ]},
                {'selector': 'table', 'props': [
                    ('border-collapse', 'collapse'),
                    ('border-radius', '12px'),
                    ('overflow', 'hidden'),
                    ('box-shadow', '0 4px 20px rgba(0,0,0,0.1)'),
                    ('width', '100%')
                ]}
            ])
            
            # 包裹在可滚动容器中，适配移动端
            table_html = f"""
            <div style="overflow-x: auto; -webkit-overflow-scrolling: touch; margin: 0 -0.5rem; padding: 0 0.5rem;">
                {styled_table.to_html()}
            </div>
            """
            st.write(table_html, unsafe_allow_html=True)
    
    # 折叠面板：详细信息
    with st.expander("📊 详细统计"):
        # 空岗情况
        vacancy_data = []
        rules = st.session_state.rules
        special_rules = rules.get("special_rules", {})
        monday_no_early_early = special_rules.get("monday_no_early_early_shift", False)
        
        for date_str in dates:
            date_schedule = schedule[date_str]
            assignments = date_schedule.get("assignments", {})
            weekday_chinese = get_weekday_chinese(date_str)
            is_monday = weekday_chinese == "周一"
            
            for shift_id, shift in shifts.items():
                if is_monday and monday_no_early_early and "早早" in shift_id:
                    continue
                required_staff = get_required_staff_for_view(shift_id, date_str, schedule, shifts)
                actual_staff = sum(1 for s in assignments.values() if s == shift_id)
                
                # 计算该班次在固定休息日有多少可用员工
                rest_day_available = 0
                required_skills = shift.get("required_skills", [])
                for emp_id, emp in employees.items():
                    if emp.get("rest_day") == weekday_chinese:
                        emp_skills = emp.get("skills", [])
                        if not required_skills or any(skill in emp_skills for skill in required_skills):
                            rest_day_available += 1
                
                # 真正的空岗 = 需要人数 - 实际上班人数 - 固定休息日可用人数
                shortage = required_staff - actual_staff
                if shortage > 0:
                    real_shortage = max(0, shortage - rest_day_available)
                    if real_shortage > 0:
                        vacancy_data.append({
                            "日期": date_str, "星期": weekday_chinese,
                            "班次": shift_id, "缺少": real_shortage
                        })
        
        if vacancy_data:
            st.warning(f"⚠️ {len(vacancy_data)} 个空岗")
            st.dataframe(pd.DataFrame(vacancy_data), use_container_width=True, hide_index=True)
        else:
            st.success("✅ 无空岗")
    
    # 折叠面板：领班优化
    with st.expander("🔧 领班优化（减少范莲彤排班）"):
        target_employee_name = "范莲彤"
        
        # 查找目标员工ID
        target_emp_id = None
        for emp_id, emp in employees.items():
            if emp.get("name") == target_employee_name:
                target_emp_id = emp_id
                break
        
        if target_emp_id:
            # 统计当前排班情况
            target_shifts_count = 0
            target_shifts_details = []
            for date_str in dates:
                date_schedule = schedule[date_str]
                assignments = date_schedule.get("assignments", {})
                if target_emp_id in assignments:
                    shift_id = assignments[target_emp_id]
                    weekday_chinese = get_weekday_chinese(date_str)
                    target_shifts_count += 1
                    target_shifts_details.append({
                        "日期": date_str,
                        "星期": weekday_chinese,
                        "班次": shift_id
                    })
            
            col1, col2 = st.columns([2, 1])
            with col1:
                st.info(f"👤 {target_employee_name}（领班）当前排班：{target_shifts_count} 天")
            with col2:
                if st.button("🔧 优化范莲彤", type="primary", use_container_width=True):
                    with st.spinner("正在优化排班..."):
                        optimized_count = 0
                        optimization_details = []
                        
                        # 按日期遍历，尝试替换范莲彤的班次
                        for date_str in dates:
                            date_schedule = schedule[date_str]
                            assignments = date_schedule.get("assignments", {})
                            
                            if target_emp_id not in assignments:
                                continue
                            
                            shift_id = assignments[target_emp_id]
                            shift = shifts.get(shift_id, {})
                            required_staff = get_required_staff_for_view(shift_id, date_str, schedule, shifts)
                            current_staff_count = sum(1 for s in assignments.values() if s == shift_id)
                            
                            if current_staff_count > required_staff:
                                del assignments[target_emp_id]
                                date_schedule["assignments"] = assignments
                                date_schedule["shift_counts"][shift_id] = current_staff_count - 1
                                optimized_count += 1
                                weekday_chinese = get_weekday_chinese(date_str)
                                optimization_details.append({
                                    "日期": date_str, "星期": weekday_chinese,
                                    "班次": shift_id, "操作": "直接移除（人数充足）"
                                })
                                continue
                            
                            replacement_found = False
                            for other_emp_id, other_emp in employees.items():
                                if other_emp_id == target_emp_id or other_emp_id in assignments:
                                    continue
                                conflicts = check_conflicts(other_emp_id, shift_id, date_str, schedule)
                                if not conflicts:
                                    del assignments[target_emp_id]
                                    assignments[other_emp_id] = shift_id
                                    date_schedule["assignments"] = assignments
                                    optimized_count += 1
                                    weekday_chinese = get_weekday_chinese(date_str)
                                    optimization_details.append({
                                        "日期": date_str, "星期": weekday_chinese,
                                        "班次": shift_id, "操作": f"替换为 {other_emp.get('name', other_emp_id)}"
                                    })
                                    replacement_found = True
                                    break
                            
                            if not replacement_found:
                                weekday_chinese = get_weekday_chinese(date_str)
                                optimization_details.append({
                                    "日期": date_str, "星期": weekday_chinese,
                                    "班次": shift_id, "操作": "无法替换（无合适人选）"
                                })
                        
                        save_json(SCHEDULE_FILE, schedule)
                        st.session_state.schedule = schedule
                        
                        if optimized_count > 0:
                            st.success(f"✅ 优化完成！已移除 {target_employee_name}（领班）{optimized_count} 个班次")
                            if optimization_details:
                                st.dataframe(pd.DataFrame(optimization_details), use_container_width=True, hide_index=True)
                            st.rerun()
                        else:
                            st.warning(f"⚠️ 无法优化：所有班次都无法移除或替换")
                            if optimization_details:
                                st.dataframe(pd.DataFrame(optimization_details), use_container_width=True, hide_index=True)
        else:
            st.info("💡 未找到员工「范莲彤」")


def analyze_schedule():
    """排班分析"""
    st.header("📊 排班分析")
    
    schedule = st.session_state.schedule
    employees = st.session_state.employees
    shifts = st.session_state.shifts
    
    if not schedule:
        st.markdown("""
        <div style="text-align: center; padding: 60px 20px; background: #f8f9fa; border-radius: 12px;">
            <h2 style="color: #6c757d;">📭 暂无排班数据</h2>
            <p style="color: #adb5bd;">请先在「🚀 生成排班」页面生成排班表</p>
        </div>
        """, unsafe_allow_html=True)
        return
    
    dates = sorted(schedule.keys())
    
    # 计算统计数据
    total_days = len(dates)
    total_assignments = sum(len(date_schedule.get("assignments", {})) for date_schedule in schedule.values())
    total_vacancies = 0
    total_required = 0
    
    # 获取特殊规则
    rules = st.session_state.rules
    special_rules = rules.get("special_rules", {})
    monday_no_early_early = special_rules.get("monday_no_early_early_shift", False)
    fixed_early_early_shift = "二期水吧-早早班"
    
    for date_str, date_schedule in schedule.items():
        assignments = date_schedule.get("assignments", {})
        weekday_chinese = get_weekday_chinese(date_str)
        is_monday = weekday_chinese == "周一"
        
        for shift_id, shift in shifts.items():
            # 如果是周一且不需要早早班，跳过早早班的空岗统计
            if is_monday and monday_no_early_early and shift_id == fixed_early_early_shift:
                continue  # 周一不需要早早班，不算空岗
            
            required_staff = get_required_staff_for_view(shift_id, date_str, schedule, shifts)  # 使用动态规则
            actual_staff = sum(1 for s in assignments.values() if s == shift_id)
            total_required += required_staff
            if actual_staff < required_staff:
                total_vacancies += (required_staff - actual_staff)
    
    with col1:
        st.metric("总天数", total_days)
    with col2:
        st.metric("总排班次数", total_assignments)
    with col3:
        st.metric("空岗数", total_vacancies)
    with col4:
        coverage_rate = round((total_assignments / total_required * 100) if total_required > 0 else 0, 1)
        st.metric("覆盖率", f"{coverage_rate}%")
    
    # 员工排班天数和工作小时统计
    st.subheader("👥 员工排班统计")
    
    employee_stats = {}
    for date_str, date_schedule in schedule.items():
        for emp_id, shift_id in date_schedule.get("assignments", {}).items():
            if emp_id not in employee_stats:
                employee_stats[emp_id] = {
                    "days": 0,
                    "hours": 0,
                    "shifts": {}
                }
            employee_stats[emp_id]["days"] += 1
            shift = shifts.get(shift_id, {})
            duration_hours = shift.get("duration_hours", 8)
            employee_stats[emp_id]["hours"] += duration_hours
            if shift_id not in employee_stats[emp_id]["shifts"]:
                employee_stats[emp_id]["shifts"][shift_id] = 0
            employee_stats[emp_id]["shifts"][shift_id] += 1
    
    if employee_stats:
        weeks = len(dates) / 7
        stats_data = []
        for emp_id, stats in employee_stats.items():
            emp = employees.get(emp_id, {})
            weekly_hours_target = emp.get("weekly_hours", 40)
            actual_weekly_hours = stats["hours"] / weeks if weeks > 0 else 0
            
            stats_data.append({
                "员工": emp.get("name", emp_id),
                "排班天数": stats["days"],
                "总工作小时": round(stats["hours"], 1),
                "平均每周天数": round(stats["days"] / weeks, 2) if weeks > 0 else 0,
                "平均每周小时": round(actual_weekly_hours, 1),
                "目标每周小时": weekly_hours_target,
                "完成度": round((actual_weekly_hours / weekly_hours_target * 100) if weekly_hours_target > 0 else 0, 1)
            })
        
        stats_df = pd.DataFrame(stats_data)
        stats_df = stats_df.sort_values("总工作小时", ascending=False)
        st.dataframe(stats_df, use_container_width=True, hide_index=True)
        
        # 可视化 - 排班天数
        col1, col2 = st.columns(2)
        with col1:
            fig = px.bar(
                stats_df,
                x="员工",
                y="排班天数",
                title="员工排班天数分布",
                labels={"员工": "员工", "排班天数": "排班天数"}
            )
        fig.update_xaxes(tickangle=45)
        st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            fig = px.bar(
                stats_df,
                x="员工",
                y="总工作小时",
                title="员工工作小时分布",
                labels={"员工": "员工", "总工作小时": "总工作小时"}
            )
            fig.update_xaxes(tickangle=45)
            st.plotly_chart(fig, use_container_width=True)
        
        # 工作强度分析
        st.markdown("#### 💪 工作强度分析")
        intensity_df = stats_df.copy()
        intensity_df["强度等级"] = intensity_df["完成度"].apply(
            lambda x: "过高" if x > 110 else ("适中" if 90 <= x <= 110 else "不足")
        )
        st.dataframe(intensity_df[["员工", "平均每周小时", "目标每周小时", "完成度", "强度等级"]], 
                    use_container_width=True, hide_index=True)
        
        # 导出分析结果
        col_csv, col_excel = st.columns(2)
        with col_csv:
            csv = stats_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                label="📥 导出为 CSV",
                data=csv,
                file_name=f"员工排班统计_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )
        with col_excel:
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                stats_df.to_excel(writer, index=False, sheet_name='排班统计')
                intensity_df[["员工", "平均每周小时", "目标每周小时", "完成度", "强度等级"]].to_excel(
                    writer, index=False, sheet_name='工作强度分析'
                )
            excel_data = output.getvalue()
            st.download_button(
                label="📥 导出为 Excel",
                data=excel_data,
                file_name=f"员工排班统计_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    # 班次使用统计
    st.subheader("⏰ 班次使用统计")
    
    shift_stats = {}
    for date_str, date_schedule in schedule.items():
        assignments = date_schedule.get("assignments", {})
        for shift_id in shifts.keys():
            if shift_id not in shift_stats:
                shift_stats[shift_id] = {
                    "count": 0,
                    "total_required": 0,
                    "total_actual": 0,
                    "vacancy_days": 0
                }
            
            required_staff = get_required_staff_for_view(shift_id, date_str, schedule, shifts)  # 使用动态规则
            actual_staff = sum(1 for s in assignments.values() if s == shift_id)
            
            shift_stats[shift_id]["count"] += 1
            shift_stats[shift_id]["total_required"] += required_staff
            shift_stats[shift_id]["total_actual"] += actual_staff
            if actual_staff < required_staff:
                shift_stats[shift_id]["vacancy_days"] += 1
    
    if shift_stats:
        usage_data = []
        for shift_id, stats in shift_stats.items():
            shift = shifts.get(shift_id, {})
            coverage_rate = round((stats["total_actual"] / stats["total_required"] * 100) if stats["total_required"] > 0 else 0, 1)
            usage_data.append({
                "班次": shift_id,
                "出现天数": stats["count"],
                "需要人数总和": stats["total_required"],
                "实际人数总和": stats["total_actual"],
                "空岗天数": stats["vacancy_days"],
                "覆盖率": f"{coverage_rate}%"
            })
        
        usage_df = pd.DataFrame(usage_data)
        usage_df = usage_df.sort_values("出现天数", ascending=False)
        st.dataframe(usage_df, use_container_width=True, hide_index=True)
        
        # 可视化
        col1, col2 = st.columns(2)
        with col1:
            fig = px.pie(
                usage_df,
                values="出现天数",
                names="班次",
                title="班次出现天数分布"
            )
        st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            fig = px.bar(
                usage_df,
                x="班次",
                y="空岗天数",
                title="班次空岗天数",
                labels={"班次": "班次", "空岗天数": "空岗天数"}
            )
            fig.update_xaxes(tickangle=45)
            st.plotly_chart(fig, use_container_width=True)
    
    # 空岗详细分析
    st.subheader("⚠️ 空岗详细分析")
    vacancy_analysis = {}
    
    # 获取特殊规则
    rules = st.session_state.rules
    special_rules = rules.get("special_rules", {})
    monday_no_early_early = special_rules.get("monday_no_early_early_shift", False)
    fixed_early_early_shift = "二期水吧-早早班"
    
    for date_str, date_schedule in schedule.items():
        assignments = date_schedule.get("assignments", {})
        weekday_chinese = get_weekday_chinese(date_str)
        is_monday = weekday_chinese == "周一"
        
        for shift_id, shift in shifts.items():
            # 如果是周一且不需要早早班，跳过早早班的空岗统计
            if is_monday and monday_no_early_early and shift_id == fixed_early_early_shift:
                continue  # 周一不需要早早班，不算空岗
            
            required_staff = get_required_staff_for_view(shift_id, date_str, schedule, shifts)  # 使用动态规则
            actual_staff = sum(1 for s in assignments.values() if s == shift_id)
            shortage = required_staff - actual_staff
            
            if shortage > 0:
                if shift_id not in vacancy_analysis:
                    vacancy_analysis[shift_id] = {
                        "total_shortage": 0,
                        "days_count": 0,
                        "details": []
                    }
                vacancy_analysis[shift_id]["total_shortage"] += shortage
                vacancy_analysis[shift_id]["days_count"] += 1
                vacancy_analysis[shift_id]["details"].append({
                    "日期": date_str,
                    "星期": weekday_chinese,
                    "缺少人数": shortage
                })
    
    if vacancy_analysis:
        vacancy_summary = []
        for shift_id, analysis in vacancy_analysis.items():
            vacancy_summary.append({
                "班次": shift_id,
                "空岗天数": analysis["days_count"],
                "累计缺少人数": analysis["total_shortage"],
                "平均每天缺少": round(analysis["total_shortage"] / analysis["days_count"], 1) if analysis["days_count"] > 0 else 0
            })
        
        vacancy_summary_df = pd.DataFrame(vacancy_summary)
        vacancy_summary_df = vacancy_summary_df.sort_values("累计缺少人数", ascending=False)
        st.dataframe(vacancy_summary_df, use_container_width=True, hide_index=True)
        
        # 空岗趋势图
        if len(dates) > 1:
            daily_vacancy = {}
            for date_str in dates:
                date_schedule = schedule[date_str]
                assignments = date_schedule.get("assignments", {})
                weekday_chinese = get_weekday_chinese(date_str)
                is_monday = weekday_chinese == "周一"
                daily_total = 0
                for shift_id, shift in shifts.items():
                    # 如果是周一且不需要早早班，跳过早早班的空岗统计
                    if is_monday and monday_no_early_early and shift_id == fixed_early_early_shift:
                        continue  # 周一不需要早早班，不算空岗
                    
                    required_staff = get_required_staff_for_view(shift_id, date_str, schedule, shifts)  # 使用动态规则
                    actual_staff = sum(1 for s in assignments.values() if s == shift_id)
                    shortage = max(0, required_staff - actual_staff)
                    daily_total += shortage
                daily_vacancy[date_str] = daily_total
            
            vacancy_trend_df = pd.DataFrame([
                {"日期": date_str, "空岗数": count}
                for date_str, count in sorted(daily_vacancy.items())
            ])
            
            fig = px.line(
                vacancy_trend_df,
                x="日期",
                y="空岗数",
                title="空岗趋势图",
                markers=True
            )
            fig.update_xaxes(tickangle=45)
            st.plotly_chart(fig, use_container_width=True)
    else:
        st.success("✅ 没有空岗情况")


def main():
    """主函数"""
    init_session_state()
    
    # 侧边栏Logo和标题
    st.sidebar.markdown("""
    <div style="text-align: center; padding: 25px 0 15px 0;">
        <div style="font-size: 48px; margin-bottom: 5px;">📅</div>
        <h3 style="color: white; margin: 5px 0; font-weight: 700; font-size: 20px;">智能排班系统</h3>
        <p style="color: rgba(255,255,255,0.6); font-size: 12px; margin: 0;">Smart Scheduling System</p>
    </div>
    """, unsafe_allow_html=True)
    
    # 显示数据状态指示器
    emp_count = len(st.session_state.employees)
    shift_count = len(st.session_state.shifts)
    schedule_count = len(st.session_state.schedule)
    
    status_color = "#06d6a0" if (emp_count > 0 and shift_count > 0) else "#ffd166"
    st.sidebar.markdown(f"""
    <div style="background: rgba(255,255,255,0.08); border-radius: 10px; padding: 12px; margin: 0 5px 15px 5px;">
        <div style="display: flex; justify-content: space-around; text-align: center;">
            <div>
                <div style="color: white; font-size: 18px; font-weight: 700;">{emp_count}</div>
                <div style="color: rgba(255,255,255,0.6); font-size: 11px;">员工</div>
            </div>
            <div style="border-left: 1px solid rgba(255,255,255,0.2); height: 35px;"></div>
            <div>
                <div style="color: white; font-size: 18px; font-weight: 700;">{shift_count}</div>
                <div style="color: rgba(255,255,255,0.6); font-size: 11px;">班次</div>
            </div>
            <div style="border-left: 1px solid rgba(255,255,255,0.2); height: 35px;"></div>
            <div>
                <div style="color: white; font-size: 18px; font-weight: 700;">{schedule_count}</div>
                <div style="color: rgba(255,255,255,0.6); font-size: 11px;">排班日</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    st.sidebar.markdown("---")
    
    # 分组式导航
    st.sidebar.markdown("""
    <p style="color: rgba(255,255,255,0.5); font-size: 11px; text-transform: uppercase; letter-spacing: 1px; margin: 10px 0 5px 5px;">
        ⚙️ 基础配置
    </p>
    """, unsafe_allow_html=True)
    
    # 侧边栏导航 - 配置模块
    config_pages = ["👥 员工管理", "⏰ 班次管理", "📐 排班规则"]
    
    st.sidebar.markdown("""
    <p style="color: rgba(255,255,255,0.5); font-size: 11px; text-transform: uppercase; letter-spacing: 1px; margin: 20px 0 5px 5px;">
        🚀 排班操作
    </p>
    """, unsafe_allow_html=True)
    
    action_pages = ["🎯 生成排班", "📋 查看排班", "📊 数据分析"]
    
    st.sidebar.markdown("""
    <p style="color: rgba(255,255,255,0.5); font-size: 11px; text-transform: uppercase; letter-spacing: 1px; margin: 20px 0 5px 5px;">
        🤖 智能工具
    </p>
    """, unsafe_allow_html=True)
    
    ai_pages = ["✨ AI 智能微调"]
    
    all_pages = config_pages + action_pages + ai_pages
    
    page = st.sidebar.radio(
        "导航菜单",
        all_pages,
        label_visibility="collapsed"
    )
    
    # 路由到不同页面
    if page == "👥 员工管理":
        employee_management()
    elif page == "⏰ 班次管理":
        shift_management()
    elif page == "📐 排班规则":
        rules_management()
    elif page == "🎯 生成排班":
        generate_schedule()
    elif page == "📋 查看排班":
        view_schedule()
    elif page == "📊 数据分析":
        analyze_schedule()
    elif page == "✨ AI 智能微调":
        ai_schedule_tuning()
    
    # AI 配置界面
    st.sidebar.markdown("---")
    
    # 显示 AI 状态指示器
    api_key = st.session_state.get("ai_api_key", "")
    if api_key:
        st.sidebar.markdown("""
        <div style="background: linear-gradient(135deg, #06d6a0 0%, #1b9aaa 100%); 
                    padding: 8px 12px; border-radius: 8px; margin: 5px;">
            <span style="color: white; font-size: 13px; font-weight: 500;">🟢 AI 已就绪</span>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.sidebar.markdown("""
        <div style="background: linear-gradient(135deg, #ffd166 0%, #ef476f 100%); 
                    padding: 8px 12px; border-radius: 8px; margin: 5px;">
            <span style="color: white; font-size: 13px; font-weight: 500;">🟡 AI 未配置</span>
        </div>
        """, unsafe_allow_html=True)
    
    with st.sidebar.expander("🔧 AI 配置", expanded=not api_key):
        # API Key 输入
        new_api_key = st.text_input(
            "🔑 API Key", 
            type="password", 
            value=st.session_state.get("ai_api_key", ""),
            help="请输入您的 Google AI Studio API Key",
            key="ai_api_key_input",
            placeholder="输入您的 API Key..."
        )
        if new_api_key != st.session_state.get("ai_api_key", ""):
            st.session_state.ai_api_key = new_api_key
        
        # 模型选择下拉框
        model_options = {
            "gemini-2.0-flash": "Gemini 2.0 Flash (推荐)",
            "gemini-1.5-pro": "Gemini 1.5 Pro (高精度)",
            "gemini-1.5-flash": "Gemini 1.5 Flash (快速)",
            "gemini-1.5-flash-8b": "Gemini 1.5 Flash 8B (轻量)",
            "custom": "自定义模型..."
        }
        
        current_model = st.session_state.get("ai_model", "gemini-2.0-flash")
        # 检查当前模型是否在预设列表中
        if current_model not in model_options and current_model != "custom":
            selected_option = "custom"
        else:
            selected_option = current_model if current_model in model_options else "gemini-2.0-flash"
        
        selected_model = st.selectbox(
            "🤖 AI 模型",
            options=list(model_options.keys()),
            format_func=lambda x: model_options[x],
            index=list(model_options.keys()).index(selected_option) if selected_option in model_options else 0,
            key="ai_model_select"
        )
        
        if selected_model == "custom":
            custom_model = st.text_input(
                "自定义模型名称",
                value=st.session_state.get("ai_model", ""),
                key="ai_model_custom_input",
                placeholder="输入模型名称..."
            )
            if custom_model:
                st.session_state.ai_model = custom_model
        else:
            st.session_state.ai_model = selected_model
        
        # 高级设置（可折叠）
        with st.expander("⚙️ 高级设置"):
            # Base URL 输入
            new_base_url = st.text_input(
                "API 地址", 
                value=st.session_state.get("ai_base_url", "https://generativelanguage.googleapis.com/v1beta/openai/"),
                help="Google Gemini 的 OpenAI 兼容接口地址",
                key="ai_base_url_input"
            )
            if new_base_url != st.session_state.get("ai_base_url", ""):
                st.session_state.ai_base_url = new_base_url
        
        # 测试连接按钮
        if st.button("🔗 测试连接", use_container_width=True, key="test_ai_connection", type="primary"):
            if not st.session_state.get("ai_api_key"):
                st.error("❌ 请先填写 API Key")
            elif not HAS_OPENAI:
                st.error("❌ 未安装 openai 库")
                st.code("pip install openai", language="bash")
            else:
                with st.spinner("正在测试连接..."):
                    try:
                        client = OpenAI(
                            api_key=st.session_state.get("ai_api_key"),
                            base_url=st.session_state.get("ai_base_url", "https://generativelanguage.googleapis.com/v1beta/openai/")
                        )
                        response = client.chat.completions.create(
                            model=st.session_state.get("ai_model", "gemini-2.0-flash"),
                            messages=[{"role": "user", "content": "你好，请回复'连接成功'"}],
                            max_tokens=20
                        )
                        st.success(f"✅ 连接成功！")
                        st.caption(f"模型: {st.session_state.get('ai_model')}")
                    except Exception as e:
                        st.error(f"❌ 连接失败")
                        st.caption(str(e)[:100])
    
    # 侧边栏底部信息
    st.sidebar.markdown("---")
    st.sidebar.markdown("""
    <div style="text-align: center; padding: 10px; opacity: 0.7;">
        <p style="font-size: 11px; color: #888; margin: 0;">
            💡 智能排班 · 高效协同
        </p>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
